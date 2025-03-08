import openpyxl
import sqlite3
import re

from bot.config import TIMEOUT_DELAY, DB_DIRECTORY, DB_TASK_DIRECTORY, DB_PATTERNS_DIRECTORY, \
    DB_MAIN_ACCOUNTS_DIRECTORY, DB_MULTI_ACCOUNTS_DIRECTORY, DB_LINKS_DIRECTORY, DB_IMAGES_DIRECTORY, DB_XLSX_DIRECTORY
from bot.handlers.commands.logging import get_task_logger, log

from typing import List

class CommandsManager:

    @staticmethod
    async def add_patterns(*, pattern_name: str, pattern: str) -> None:
        """Добавление шаблона"""

        try:
            with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute("INSERT INTO Patterns (pattern_name, pattern) VALUES (?, ?)",
                               (pattern_name, pattern,))
                connection.commit()
        except Exception as e:
            log.debug(f"Произошла ошибка при добавлении шаблона: {str(e)}")

    @staticmethod
    async def add_link(*, link_name: str, link_source: str) -> None:
        """Добавление ссылки"""

        try:
            with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute("INSERT INTO Links (link_name, link_source) VALUES (?, ?)",
                               (link_name, link_source,))
                connection.commit()
        except Exception as e:
            log.debug(f"Произошла ошибка при добавлении ссылки: {str(e)}")

    @staticmethod
    async def remove_patterns(*, pattern_id: int) -> None:
        """Удаление шаблона"""

        try:
            with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute("DELETE FROM Patterns WHERE id = ?", (pattern_id,))
                connection.commit()
        except Exception as e:
            log.debug(f"Произошла ошибка при удалении шаблона: {str(e)}")

    @staticmethod
    async def remove_link(*, link_id: int) -> None:
        """Удаление ссылки"""

        try:
            with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute("DELETE FROM Links WHERE id = ?", (link_id,))
                connection.commit()
        except Exception as e:
            log.debug(f"Произошла ошибка при удалении ссылки: {str(e)}")

    @staticmethod
    async def add_keys_to_db(keys):
        keys_str = '\n'.join(keys)

        with sqlite3.connect(DB_XLSX_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
            cursor = connection.cursor()

            # Проверяем, существует ли уже такая запись
            cursor.execute("SELECT id FROM Xlsx WHERE keys = ?", (keys_str,))
            row = cursor.fetchone()

            if row:
                return row[0]  # Возвращаем id уже существующей записи

            # Добавляем новую запись
            cursor.execute("INSERT INTO Xlsx (keys) VALUES (?)", (keys_str,))
            connection.commit()

            return cursor.lastrowid  # Возвращаем id только что добавленной записи

    @staticmethod
    async def generate_prompts_from_xlsx(task: str, file_path: str) -> tuple:
        """Получение тем и ключей промтов из xlsx-файла"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            prompts = []
            current_topic = None
            current_pattern_name = None
            current_link_name = None
            current_image_name = None
            theme_count = 0
            current_keys = []
            header_key = None  # Сохраняем ключ заголовка, если опциональная тема не задана
            is_optional = False  # Флаг, указывающий, что тема опциональная

            with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute("SELECT pattern_name, pattern FROM Patterns")
                patterns = cursor.fetchall()
                connection.commit()

            for row in sheet.iter_rows(min_row=2, values_only=False):
                key_cell = row[0]
                pattern_name_cell = row[1]
                link_name_cell = row[2]
                option_topic_cell = row[3]
                image_name_cell = row[4]

                # Если ячейка с ключом имеет изменённый фон – начинаем новую тему
                if key_cell.fill.start_color.index != '00000000':
                    # Завершаем предыдущую тему (если она была)
                    if current_topic is not None:
                        # Если нет дополнительных ключей, добавляем header_key (только для НЕопциональной темы)
                        if not current_keys and header_key is not None:
                            current_keys.append(header_key)

                        # Формирование списка ключей для БД:
                        # Если тема опциональная, не добавляем её в список ключей
                        if is_optional:
                            keys_to_send = current_keys
                        else:
                            # Если тема не опциональная, добавляем current_topic, если его там ещё нет
                            keys_to_send = current_keys if current_topic in current_keys else [
                                                                                                  current_topic] + current_keys

                        xlsx_id = await CommandsManager.add_keys_to_db(keys_to_send)
                        prompt = await CommandsManager.generate_prompt(
                            current_topic, current_pattern_name, keys_to_send,
                            patterns, task, current_link_name, current_image_name
                        )
                        if prompt is not None:
                            prompts.append((prompt, current_topic, xlsx_id))

                    # Сбрасываем список ключей для новой темы
                    current_keys = []

                    # Определяем, задана ли опциональная тема
                    if option_topic_cell.value:
                        current_topic = option_topic_cell.value  # Опциональная тема
                        # Добавляем только ключ из key_cell
                        current_keys.append(key_cell.value)
                        header_key = None  # Заголовочный ключ не нужен
                        is_optional = True
                    else:
                        # Если опциональной темы нет, используем значение ключа как тему
                        current_topic = key_cell.value
                        header_key = key_cell.value
                        is_optional = False

                    current_pattern_name = pattern_name_cell.value
                    current_link_name = link_name_cell.value
                    current_image_name = image_name_cell.value
                    theme_count += 1
                else:
                    # Для строк без изменённого фона добавляем ключ в current_keys,
                    # только если значение ключа действительно присутствует
                    if key_cell.value is not None and str(key_cell.value).strip() != "":
                        current_keys.append(key_cell.value)

            # Обработка последней темы, если она осталась
            if current_topic is not None:
                if not current_keys and header_key is not None:
                    current_keys.append(header_key)

                if is_optional:
                    keys_to_send = current_keys
                else:
                    keys_to_send = current_keys if current_topic in current_keys else [current_topic] + current_keys

                xlsx_id = await CommandsManager.add_keys_to_db(keys_to_send)
                prompt = await CommandsManager.generate_prompt(
                    current_topic, current_pattern_name, keys_to_send,
                    patterns, task, current_link_name, current_image_name
                )
                if prompt is not None:
                    prompts.append((prompt, current_topic, xlsx_id))

            workbook.close()
            return prompts, theme_count

        except Exception as e:
            task_log = get_task_logger(task)
            task_log.debug(f"Произошла ошибка при получении тем и ключей промтов: {str(e)}")
            return False, None


        except Exception as e:
            task_log = get_task_logger(task)
            task_log.debug(f"Произошла ошибка при получении тем и ключей промтов: {str(e)}")
            return False, None

    @staticmethod
    async def generate_prompt(topic: str,
                              pattern_name: str,
                              keys: List[str],
                              patterns: List[str],
                              task: str,
                              pattern_link_name: str,
                              pattern_image_name) -> str:
        """Генерация текста промта на основе шаблона"""

        connection = sqlite3.connect(DB_DIRECTORY + task + '.db', timeout=TIMEOUT_DELAY)
        cursor = connection.cursor()
        cursor.execute("SELECT count_key_words FROM TasksSettings WHERE id = 1")
        count = cursor.fetchone()[0]
        connection.close()

        if count == '-':
            with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute("SELECT count_key_words FROM TasksSettings WHERE id = 1")
                count = cursor.fetchone()[0]

        try:
            pattern = next((pattern for name, pattern in patterns if str(pattern_name) == str(name)), None)

            if pattern is None:
                task_log = get_task_logger(task)
                task_log.debug(f"Шаблон отсутствует для темы '{topic}' с pattern_name - '{pattern_name}'")
                return None

            keys_str = "\n".join(filter(None, keys[:int(count)]))

            if pattern_link_name:

                with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                    cursor = connection.cursor()
                    cursor.execute("SELECT link_name, link_source FROM Links")
                    links = cursor.fetchall()
                    connection.commit()

                flag = False

                for link_name, link_source in links:

                    if pattern_link_name == link_name:
                        flag = True
                        pattern = pattern.replace('%LINKS%', link_source)
                        break

                if not flag:
                    pattern = pattern.replace('%LINKS%', '')
            else:
                pattern = pattern.replace('%LINKS%', '')

            if pattern_image_name:

                with sqlite3.connect(DB_IMAGES_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                    cursor = connection.cursor()
                    cursor.execute("SELECT image_name, image_path FROM Images")
                    images = cursor.fetchall()
                    connection.commit()

                flag = False

                for image_name, image_path in images:
                    if pattern_image_name == image_name:
                        flag = True
                        matches = re.findall(r'/([^/]+)\.[a-zA-Z0-9]+', image_path)
                        pattern = pattern.replace('%IMAGES%',
                                                  f'В соответствующем названию подзаголовке добавь <div type="image">Name</div>, '
                                                  f'где Name соответствует названию продукта. Названия картинок:\n'
                                                  f'{'\n'.join(matches)}')
                        break

                if not flag:
                    pattern = pattern.replace('%IMAGES%', '')
            else:
                pattern = pattern.replace('%IMAGES%', '')

            return pattern.replace("%NAME%", topic).replace("%KEYS%", keys_str)

        except Exception as e:
            task_log = get_task_logger(task)
            task_log.debug(f"Произошла ошибка при генерации текста промтов: {str(e)}")
            return None

    @staticmethod
    async def save_prompts_to_db(task: str, prompts: List[str]):
        """Сохранение сгенерированных промтов в базу данных"""

        try:
            connection = sqlite3.connect(DB_DIRECTORY + task + '.db', timeout=TIMEOUT_DELAY)
            cursor = connection.cursor()
            for prompt, prompt_theme, xlsx_id in prompts:
                cursor.execute(
                    "INSERT INTO Prompts (prompt, prompt_theme, xlsx_id) VALUES (?, ?, ?)",
                    (prompt, prompt_theme, str(xlsx_id))
                )
            connection.commit()
            connection.close()
            return True
        except Exception as e:
            task_log = get_task_logger(task)
            task_log.debug(f"Произошла ошибка при сохранении сгенерированных промтов в базу данных: {str(e)}")
            return False

    @staticmethod
    async def save_priority_prompts_to_db(task: str, priority_prompts: str):
        """Сохранение диапазона приоритетных промтов в базу данных"""

        try:
            with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "UPDATE Tasks SET priority_prompts = ? WHERE task_name = ?",
                    (priority_prompts, task)
                )
                connection.commit()
            return True
        except Exception as e:
            task_log = get_task_logger(task)
            task_log.debug(f"Произошла ошибка при сохранении диапазона приоритетных промтов в базу данных: {str(e)}")
            return False

    @staticmethod
    async def save_theme_count(task: str, theme_count: str):
        """Сохранение количества распознанных тем"""

        try:
            with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "UPDATE Tasks SET theme_count = ? WHERE task_name = ?",
                    (theme_count, task)
                )
                connection.commit()
            return True
        except Exception as e:
            task_log = get_task_logger(task)
            task_log.debug(f"Произошла ошибка при сохранении количества распознанных тем в базу данных: {str(e)}")
            return False

    @staticmethod
    async def update_task_status_db(task: str, status: str):
        """Обновление статуса задания в базе данных"""

        try:
            with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "UPDATE Tasks SET status = ? WHERE task_name = ?",
                    (status, task)
                )
                connection.commit()
        except Exception as e:
            task_log = get_task_logger(task)
            task_log.debug(f"Произошла ошибка при обновлении статуса задания в базе данных: {str(e)}")

    @staticmethod
    async def save_accounts_to_db(file_path, task_type):
        """Сохраняет данные из xlsx-файла в таблицу базы данных Accounts."""

        with sqlite3.connect(DB_MAIN_ACCOUNTS_DIRECTORY if task_type == 'Основной'\
                                     else DB_MULTI_ACCOUNTS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
            cursor = connection.cursor()

            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                email, password, login, ip, port, proxy_login, proxy_password, url = row
                cursor.execute(
                    """
                    INSERT INTO Accounts (
                        account_email, account_password, account_login, 
                        proxy_ip, proxy_port, proxy_login, proxy_password, account_url
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (email, password, login, ip, port, proxy_login, proxy_password, url)
                )

            connection.commit()

    @staticmethod
    async def delete_accounts_by_ids(ids, task_type):
        """Удаляет аккаунты из базы данных по указанным ID."""

        with sqlite3.connect(DB_MAIN_ACCOUNTS_DIRECTORY if task_type == 'Основной'\
                             else DB_MULTI_ACCOUNTS_DIRECTORY) as connection:
            cursor = connection.cursor()

            cursor.execute(
                f"SELECT id FROM Accounts WHERE id IN ({','.join('?' * len(ids))})",
                ids
            )
            existing_ids = [row[0] for row in cursor.fetchall()]

            if len(existing_ids) < len(ids):
                missing_ids = set(ids) - set(existing_ids)
                return False, missing_ids

            cursor.execute(
                f"DELETE FROM Accounts WHERE id IN ({','.join('?' * len(existing_ids))})",
                existing_ids
            )
            connection.commit()

            return True, existing_ids
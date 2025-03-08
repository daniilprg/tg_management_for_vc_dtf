import os
import sqlite3
import uuid

from html import escape

import openpyxl
from aiogram import Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile

from bot.config import TIMEOUT_DELAY, DB_PATTERNS_DIRECTORY
from bot.handlers.commands.admins_filter import AdminFilter
from bot.handlers.commands.commands_manager import CommandsManager
from bot.handlers.routers.control_panel import BACK_TO_TASKS
from bot.keyboards.keyboards import patterns_list

from openpyxl import load_workbook

router_tasks_patterns = Router(name=__name__)

class PatternMessage(StatesGroup):
    message = State()
    create = State()

class PatternMoreDownload(StatesGroup):
    message = State()

class PatternNameEdit(StatesGroup):
    message = State()

class PatternSourceEdit(StatesGroup):
    message = State()

@router_tasks_patterns.callback_query(lambda call: call.data.startswith('patterns-list'), AdminFilter())
async def task_patterns_callback_query(call: CallbackQuery):
    keyboard = await patterns_list()
    await call.message.edit_text("Выберите шаблон или добавьте новый:", reply_markup=keyboard)

@router_tasks_patterns.callback_query(lambda call: call.data.startswith('self-pattern-'), AdminFilter())
async def self_pattern_callback_query(call: CallbackQuery):
    pattern_id = call.data[13:]

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='✏️ Редактировать название', callback_data=f'patt-name-edit-{pattern_id}')],
        [InlineKeyboardButton(text='✏️ Редактировать содержимое', callback_data=f'patt-source-edit-{pattern_id}')],
        [InlineKeyboardButton(text='🗑 Удалить шаблон', callback_data=f'pattern-delete-{pattern_id}')],
        [InlineKeyboardButton(text='🗃 Выгрузка шаблона', callback_data=f'pattern-download-{pattern_id}')],
        [InlineKeyboardButton(text='⬅️ Назад к шаблонам', callback_data=f'patterns-list')]
    ])

    with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT pattern FROM Patterns WHERE id = ?", (pattern_id,))
        pattern = cursor.fetchone()[0]

    if len(pattern) > 3500:
        pattern = pattern[:3500] + "..."

    await call.message.edit_text("📃 <b>Шаблон:</b>\n\n"
                                 f"{escape(pattern)}", reply_markup=keyboard)

@router_tasks_patterns.callback_query(lambda call: call.data.startswith('pattern-download-'), AdminFilter())
async def pattern_download_callback_query(call: CallbackQuery):
    pattern_id = call.data[17:]

    with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT pattern FROM Patterns WHERE id = ?", (pattern_id,))
        pattern = cursor.fetchone()[0]

    file_path = f"bot/assets/txt/pattern_{uuid.uuid4()}.txt"

    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(pattern)

    await call.message.answer_document(FSInputFile(file_path))
    os.remove(file_path)
    keyboard = await patterns_list()
    await call.message.answer("Выберите шаблон или добавьте новый:", reply_markup=keyboard)

@router_tasks_patterns.callback_query(lambda call: call.data.startswith('pattern-delete-'), AdminFilter())
async def pattern_delete_callback_query(call: CallbackQuery):
    pattern_id = call.data[15:]
    await CommandsManager.remove_patterns(pattern_id=pattern_id)
    keyboard = await patterns_list()
    await call.message.edit_text("Выберите шаблон или добавьте новый:", reply_markup=keyboard)


@router_tasks_patterns.callback_query(lambda call: call.data.startswith('create-pattern'), AdminFilter())
async def create_pattern_name_callback_query(call: CallbackQuery, state: FSMContext):
    await call.message.edit_text("💬 Отправьте название шаблона в сообщении", reply_markup=BACK_TO_TASKS)
    await state.set_state(PatternMessage.message)

@router_tasks_patterns.message(PatternMessage.message, AdminFilter())
async def create_pattern_name_handler(message: Message, state: FSMContext):
    pattern_name = message.text
    await message.answer("💬 Отправьте шаблон в сообщении или через txt-файл\n\n"
                                 "%NAME% - тема статьи\n"
                                 "%KEYS% - ключевые слова\n"
                                 "%LINKS% - группа ссылок\n"
                                 "%IMAGES% - группа изображений", reply_markup=BACK_TO_TASKS)
    await state.update_data(pattern_name=pattern_name)
    await state.set_state(PatternMessage.create)

@router_tasks_patterns.message(PatternMessage.create, AdminFilter())
async def create_pattern_handler(message: Message, state: FSMContext):
    pattern_data = await state.get_data()
    pattern_name = pattern_data.get('pattern_name')

    if message.document:
        if not message.document.file_name.endswith('.txt'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .txt", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with open(file_path, 'r', encoding='utf-8') as f:
            pattern_text = f.read()
        os.remove(file_path)
    elif message.text:
        pattern_text = message.text
    else:
        await message.answer("❌ Пожалуйста, отправьте текст или txt-файл с шаблоном.", reply_markup=BACK_TO_TASKS)
        return

    if "%NAME%" not in pattern_text or "%KEYS%" not in pattern_text:
        await message.answer(
            "❌ Шаблон должен содержать <b>%NAME%</b> и <b>%KEYS%</b>.\n\nПожалуйста, отправьте корректный шаблон заново.", reply_markup=BACK_TO_TASKS
        )
        return

    await CommandsManager.add_patterns(pattern_name=pattern_name, pattern=pattern_text)
    keyboard = await patterns_list()
    await message.answer('✅ Шаблон успешно добавлен!')
    await message.answer("Выберите шаблон или добавьте новый:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()


@router_tasks_patterns.callback_query(lambda call: call.data.startswith('patt-name-edit-'), AdminFilter())
async def edit_pattern_name_callback_query(call: CallbackQuery, state: FSMContext):
    pattern_id = call.data[15:]
    await call.message.edit_text("💬 Отправьте новое название шаблона в сообщении", reply_markup=BACK_TO_TASKS)
    await state.update_data(pattern_id=pattern_id)
    await state.set_state(PatternNameEdit.message)

@router_tasks_patterns.message(PatternNameEdit.message, AdminFilter())
async def edit_pattern_name_handler(message: Message, state: FSMContext):
    pattern_data = await state.get_data()
    pattern_id = pattern_data.get('pattern_id')

    with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE Patterns SET pattern_name = ? WHERE id = ?", (message.text, pattern_id)
        )
        connection.commit()

    keyboard = await patterns_list()
    await message.answer('✅ Название шаблона успешно изменено!')
    await message.answer("Выберите шаблон или добавьте новый:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_patterns.callback_query(lambda call: call.data.startswith('patt-source-edit-'), AdminFilter())
async def edit_pattern_source_callback_query(call: CallbackQuery, state: FSMContext):
    pattern_id = call.data[17:]
    await call.message.edit_text("💬 Отправьте новое содержимое шаблона в сообщении или txt-файле", reply_markup=BACK_TO_TASKS)
    await state.update_data(pattern_id=pattern_id)
    await state.set_state(PatternSourceEdit.message)

@router_tasks_patterns.message(PatternSourceEdit.message, AdminFilter())
async def edit_pattern_source_handler(message: Message, state: FSMContext):
    pattern_data = await state.get_data()
    pattern_id = pattern_data.get('pattern_id')

    if message.document:
        if not message.document.file_name.endswith('.txt'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .txt", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with open(file_path, 'r', encoding='utf-8') as f:
            pattern_source = f.read()
        os.remove(file_path)
    elif message.text:
        pattern_source = message.text
    else:
        await message.answer("❌ Пожалуйста, отправьте текст или txt-файл с содержимым шаблона.", reply_markup=BACK_TO_TASKS)
        return

    if "%NAME%" not in pattern_source or "%KEYS%" not in pattern_source:
        await message.answer(
            "❌ Шаблон должен содержать <b>%NAME%</b> и <b>%KEYS%</b>.\n\nПожалуйста, отправьте корректный шаблон заново.", reply_markup=BACK_TO_TASKS
        )
        return

    with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE Patterns SET pattern = ? WHERE id = ?", (pattern_source, pattern_id)
        )
        connection.commit()

    keyboard = await patterns_list()
    await message.answer('✅ Содержимое шаблона успешно изменено!')
    await message.answer("Выберите шаблон или добавьте новый:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_patterns.callback_query(lambda call: call.data.startswith('download-patterns'), AdminFilter())
async def pattern_download_callback_query(call: CallbackQuery):

    with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM Patterns")
        file_path = f"bot/assets/xlsx/patterns_{uuid.uuid4()}.xlsx"
        links_data = cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Patterns"
        ws.append(["ID", "Pattern_name", "Pattern"])
        for row in links_data:
            ws.append(row)
        wb.save(file_path)

    await call.message.answer_document(FSInputFile(file_path),
                                       caption="Конвертированная таблица базы данных в xlsx-файл.")
    os.remove(file_path)
    keyboard = await patterns_list()
    await call.message.answer("Выберите шаблон или добавьте новый:", reply_markup=keyboard)

@router_tasks_patterns.callback_query(lambda call: call.data.startswith('all-create-patterns'), AdminFilter())
async def all_create_patterns_callback_query(call: CallbackQuery, state: FSMContext):
    await call.message.edit_text("💬 Отправьте xlsx-файл с названием шаблонов и их содержимым.\n\n"
                                 "<b>Формат файла:</b>\n"
                                 "Название | Шаблон\n\n"
                                 "<b>Примечание:</b> первая строка считается как название столбцов, данные должны идти со второй строки.", reply_markup=BACK_TO_TASKS
                                 )
    await state.set_state(PatternMoreDownload.message)

@router_tasks_patterns.message(PatternMoreDownload.message, AdminFilter())
async def all_create_patterns_handler(message: Message, state: FSMContext):
    if message.document:
        if not message.document.file_name.endswith('.xlsx'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .xlsx", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/xlsx/patterns_{uuid.uuid4()}.xlsx"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
            cursor = connection.cursor()
            wb = load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                pattern_name, pattern = row
                if pattern_name and pattern:
                    if "%NAME%" not in pattern or "%KEYS%" not in pattern:
                        await message.answer(
                            f"❌ Шаблон ({pattern_name}) не будет загружен/отредактирован, так как он должен содержать <b>%NAME%</b> и <b>%KEYS%</b>.", reply_markup=BACK_TO_TASKS
                        )
                    else:
                        cursor.execute("SELECT COUNT(*) FROM Patterns WHERE pattern_name = ?", (pattern_name,))
                        exists = cursor.fetchone()[0] > 0

                        if exists:
                            cursor.execute("UPDATE Patterns SET pattern = ? WHERE pattern_name = ?", (pattern, pattern_name))
                        else:
                            cursor.execute("INSERT INTO Patterns (pattern_name, pattern) VALUES (?, ?)",
                                           (pattern_name, pattern))

            connection.commit()

    else:
        await message.answer("❌ Пожалуйста, отправьте xlsx-файл с названием шаблонов и их содержимым.", reply_markup=BACK_TO_TASKS)
        return

    keyboard = await patterns_list()
    await message.answer('✅ Шаблоны успешно выгружены и добавлены!')
    await message.answer("Выберите шаблон или добавьте новый:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()
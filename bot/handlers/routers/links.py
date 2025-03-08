import os
import sqlite3
import uuid

from html import escape

import openpyxl
from aiogram import Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile

from bot.config import TIMEOUT_DELAY, DB_LINKS_DIRECTORY
from bot.handlers.commands.admins_filter import AdminFilter
from bot.handlers.commands.commands_manager import CommandsManager
from bot.handlers.routers.control_panel import BACK_TO_TASKS
from bot.keyboards.keyboards import links_list

from openpyxl import load_workbook

router_tasks_links = Router(name=__name__)

class LinkMessage(StatesGroup):
    link_name = State()
    message = State()

class LinksMessage(StatesGroup):
    message = State()

class LinkNameEdit(StatesGroup):
    message = State()

class LinkSourceEdit(StatesGroup):
    message = State()

@router_tasks_links.callback_query(lambda call: call.data.startswith('links-list'), AdminFilter())
async def task_links_callback_query(call: CallbackQuery):
    keyboard = await links_list()
    await call.message.edit_text("Выберите ссылку или добавьте новую:", reply_markup=keyboard)

@router_tasks_links.callback_query(lambda call: call.data.startswith('self-links-'), AdminFilter())
async def self_links_callback_query(call: CallbackQuery):
    link_id = call.data[11:]

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='✏️ Редактировать название', callback_data=f'link-name-edit-{link_id}')],
        [InlineKeyboardButton(text='✏️ Редактировать содержимое', callback_data=f'link-source-edit-{link_id}')],
        [InlineKeyboardButton(text='🗑 Удалить ссылку', callback_data=f'link-delete-{link_id}')],
        [InlineKeyboardButton(text='🗃 Выгрузка ссылки', callback_data=f'link-download-{link_id}')],
        [InlineKeyboardButton(text='⬅️ Назад к ссылкам', callback_data=f'links-list')]
    ])

    with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT link_name, link_source FROM Links WHERE id = ?", (link_id,))
        link_name, link_source = cursor.fetchone()

    if len(link_source) > 3500:
        link_source = link_source[:3500] + "..."

    await call.message.edit_text(f"📃 <b>Ссылка ({link_name}):</b>\n\n"
                                 f"{escape(link_source)}", reply_markup=keyboard)


@router_tasks_links.callback_query(lambda call: call.data.startswith('link-name-edit-'), AdminFilter())
async def edit_link_name_callback_query(call: CallbackQuery, state: FSMContext):
    link_id = call.data[15:]
    await call.message.edit_text("💬 Отправьте новое название ссылки в сообщении", reply_markup=BACK_TO_TASKS)
    await state.update_data(link_id=link_id)
    await state.set_state(LinkNameEdit.message)

@router_tasks_links.message(LinkNameEdit.message, AdminFilter())
async def edit_link_name_handler(message: Message, state: FSMContext):
    link_data = await state.get_data()
    link_id = link_data.get('link_id')

    with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE Links SET link_name = ? WHERE id = ?", (message.text, link_id)
        )
        connection.commit()

    keyboard = await links_list()
    await message.answer('✅ Название ссылки успешно изменено!')
    await message.answer("Выберите ссылку или добавьте новую:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_links.callback_query(lambda call: call.data.startswith('link-source-edit-'), AdminFilter())
async def edit_link_source_callback_query(call: CallbackQuery, state: FSMContext):
    link_id = call.data[17:]
    await call.message.edit_text("💬 Отправьте новое содержимое ссылки в сообщении или txt-файле", reply_markup=BACK_TO_TASKS)
    await state.update_data(link_id=link_id)
    await state.set_state(LinkSourceEdit.message)

@router_tasks_links.message(LinkSourceEdit.message, AdminFilter())
async def edit_link_source_handler(message: Message, state: FSMContext):
    link_data = await state.get_data()
    link_id = link_data.get('link_id')

    if message.document:
        if not message.document.file_name.endswith('.txt'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .txt", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with open(file_path, 'r', encoding='utf-8') as f:
            link_source = f.read()
        os.remove(file_path)
    elif message.text:
        link_source = message.text
    else:
        await message.answer("❌ Пожалуйста, отправьте текст или txt-файл с блоком ссылки.", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE Links SET link_source = ? WHERE id = ?", (link_source, link_id)
        )
        connection.commit()

    keyboard = await links_list()
    await message.answer('✅ Содержимое ссылки успешно изменено!')
    await message.answer("Выберите ссылку или добавьте новую:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_links.callback_query(lambda call: call.data.startswith('link-download-'), AdminFilter())
async def link_download_callback_query(call: CallbackQuery):
    link_id = call.data[14:]

    with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT link_name, link_source FROM Links WHERE id = ?", (link_id,))
        link_name, link_source = cursor.fetchone()

    file_path = f"bot/assets/txt/link_{link_name}.txt"

    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(link_source)

    await call.message.answer_document(FSInputFile(file_path))
    os.remove(file_path)
    keyboard = await links_list()
    await call.message.answer("Выберите ссылку или добавьте новую:", reply_markup=keyboard)

@router_tasks_links.callback_query(lambda call: call.data.startswith('link-delete-'), AdminFilter())
async def link_delete_callback_query(call: CallbackQuery):
    link_id = call.data[12:]
    await CommandsManager.remove_link(link_id=link_id)
    keyboard = await links_list()
    await call.message.edit_text("Выберите ссылку или добавьте новую:", reply_markup=keyboard)

@router_tasks_links.callback_query(lambda call: call.data.startswith('all-create-links'), AdminFilter())
async def all_create_link_name_callback_query(call: CallbackQuery, state: FSMContext):
    await call.message.edit_text("💬 Отправьте xlsx-файл с названием ссылок и их содержимым.\n"
                                 "<b>Формат файла:</b>\n"
                                 "Название | Ссылки\n\n"
                                 "<b>Примечание:</b> первая строка считается как название столбцов, данные должны идти со второй строки.", reply_markup=BACK_TO_TASKS
                                 )
    await state.set_state(LinksMessage.message)

@router_tasks_links.message(LinksMessage.message, AdminFilter())
async def all_create_link_name_handler(message: Message, state: FSMContext):
    if message.document:
        if not message.document.file_name.endswith('.xlsx'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .xlsx", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/xlsx/links_{uuid.uuid4()}.xlsx"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
            cursor = connection.cursor()
            wb = load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                link_name, link_source = row
                if link_name and link_source:
                    cursor.execute("SELECT COUNT(*) FROM Links WHERE link_name = ?", (link_name,))
                    exists = cursor.fetchone()[0] > 0

                    if exists:
                        cursor.execute("UPDATE Links SET link_source = ? WHERE link_name = ?", (link_source, link_name))
                    else:
                        cursor.execute("INSERT INTO Links (link_name, link_source) VALUES (?, ?)",
                                       (link_name, link_source))

            connection.commit()

    else:
        await message.answer("❌ Пожалуйста, отправьте xlsx-файл с названием ссылок и их содержимым.", reply_markup=BACK_TO_TASKS)
        return

    keyboard = await links_list()
    await message.answer('✅ Ссылки успешно выгружены и добавлены!')
    await message.answer("Выберите ссылку или добавьте новую:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_links.callback_query(lambda call: call.data.startswith('create-link'), AdminFilter())
async def create_link_name_callback_query(call: CallbackQuery, state: FSMContext):
    await call.message.edit_text("💬 Отправьте название ссылки в сообщении", reply_markup=BACK_TO_TASKS)
    await state.set_state(LinkMessage.link_name)

@router_tasks_links.message(LinkMessage.link_name, AdminFilter())
async def create_link_name_handler(message: Message, state: FSMContext):
    link_name = message.text
    await message.answer("💬 Отправьте блок ссылки в сообщении или txt-файле", reply_markup=BACK_TO_TASKS)
    await state.update_data(link_name=link_name)
    await state.set_state(LinkMessage.message)

@router_tasks_links.message(LinkMessage.message, AdminFilter())
async def create_link_handler(message: Message, state: FSMContext):
    link_data = await state.get_data()
    link_name = link_data.get('link_name')

    if message.document:
        if not message.document.file_name.endswith('.txt'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .txt", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with open(file_path, 'r', encoding='utf-8') as f:
            link_source = f.read()
        os.remove(file_path)
    elif message.text:
        link_source = message.text
    else:
        await message.answer("❌ Пожалуйста, отправьте текст или txt-файл с блоком ссылки.", reply_markup=BACK_TO_TASKS)
        return

    await CommandsManager.add_link(link_name=link_name, link_source=link_source)
    keyboard = await links_list()
    await message.answer('✅ Ссылка успешно добавлена!')
    await message.answer("Выберите ссылку или добавьте новую:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_links.callback_query(lambda call: call.data.startswith('download-links'), AdminFilter())
async def links_download_callback_query(call: CallbackQuery):

    with sqlite3.connect(DB_LINKS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM Links")
        file_path = f"bot/assets/xlsx/links_{uuid.uuid4()}.xlsx"
        links_data = cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Links"
        ws.append(["ID", "Link_name", "Link_source"])
        for row in links_data:
            ws.append(row)
        wb.save(file_path)

    await call.message.answer_document(FSInputFile(file_path),
                                       caption="Конвертированная таблица базы данных в xlsx-файл.")
    os.remove(file_path)
    keyboard = await links_list()
    await call.message.answer("Выберите ссылку или добавьте новую:", reply_markup=keyboard)
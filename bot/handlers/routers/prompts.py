import os
import re
import sqlite3
import openpyxl
import uuid

from aiogram import Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, CallbackQuery, FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton

from bot.config import DB_DIRECTORY, TIMEOUT_DELAY, DB_PATTERNS_DIRECTORY
from bot.handlers.commands.admins_filter import AdminFilter
from bot.handlers.commands.commands_manager import CommandsManager
from bot.handlers.routers.control_panel import BACK_TO_TASKS
from bot.keyboards.keyboards import task_prompts

from openpyxl import load_workbook

router_tasks_prompts = Router(name=__name__)

class PromptsMessage(StatesGroup):
    upload_xlsx = State()

class PromptMoreDownload(StatesGroup):
    message = State()

class PromptsPriority(StatesGroup):
    message = State()

class PromptsEdit(StatesGroup):
    message = State()
    txt = State()

class PromptsThemeEdit(StatesGroup):
    message = State()
    txt = State()

@router_tasks_prompts.callback_query(lambda call: call.data.startswith('prompts-list-'), AdminFilter())
async def task_prompts_callback_query(call: CallbackQuery):
    task_name = call.data[13:]
    text, keyboard = await task_prompts(task_name=task_name)
    await call.message.edit_text(text, reply_markup=keyboard)

@router_tasks_prompts.callback_query(lambda call: call.data.startswith('download-prompts-'), AdminFilter())
async def download_prompts_callback_query(call: CallbackQuery):
    task_name = call.data[17:]
    connection = sqlite3.connect(DB_DIRECTORY + task_name + '.db', timeout=TIMEOUT_DELAY)
    cursor = connection.cursor()
    cursor.execute("SELECT id, prompt, prompt_theme, marks FROM Prompts")
    prompts_data = cursor.fetchall()
    connection.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Promts_{task_name}"
    ws.append(["ID", "Promts", "Prompts_Theme", "Marks"])
    for row in prompts_data:
        ws.append(row)
    file_path = f"bot/assets/xlsx/prompts_{task_name}.xlsx"
    wb.save(file_path)
    await call.message.answer_document(FSInputFile(file_path), caption="Конвертированная таблица базы данных в xlsx-файл.")
    os.remove(file_path)
    text, keyboard = await task_prompts(task_name=task_name)
    await call.message.answer(text, reply_markup=keyboard)


@router_tasks_prompts.callback_query(lambda call: call.data.startswith('upload-xlsx-'), AdminFilter())
async def upload_xlsx_callback_query(call: CallbackQuery, state: FSMContext):
    task_name = call.data[12:]
    await call.message.edit_text('Отправьте xlsx-файл со списком ключевых фраз, шаблонов, ссылок, тем и изображений.\n\n'
                                 '<b>Формат файла:</b>\n'
                                 'Ключи | Шаблоны | Ссылки | Темы | Изображения\n\n'
                                 '<b>Примечание:</b> первая строка считается как название столбцов, данные должны идти со второй строки.', reply_markup=BACK_TO_TASKS)
    await state.update_data(task_name=task_name)
    await state.set_state(PromptsMessage.upload_xlsx)


@router_tasks_prompts.message(PromptsMessage.upload_xlsx, AdminFilter())
async def upload_xlsx_handler(message: Message, state: FSMContext):

    prompt_data = await state.get_data()
    task_name = prompt_data.get('task_name')

    with sqlite3.connect(DB_PATTERNS_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT COUNT(*) FROM Patterns")
        patterns_count = cursor.fetchone()[0]

    if patterns_count == 0:
        await message.answer("❌ Пожалуйста, добавьте шаблоны и попробуйте снова.", reply_markup=BACK_TO_TASKS)
        return

    document = message.document
    if not document.file_name.endswith('.xlsx'):
        await message.answer("❌ Пожалуйста, отправьте файл в формате xlsx.", reply_markup=BACK_TO_TASKS)
        return

    file_path = os.path.join('bot/assets/xlsx', f'keys_{task_name}.xlsx')
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    bot = message.bot
    await bot.download(document.file_id, destination=file_path)

    prompts, theme_count = await CommandsManager.generate_prompts_from_xlsx(task=task_name, file_path=file_path)

    flag_save_prompts = await CommandsManager.save_prompts_to_db(task=task_name, prompts=prompts)
    flag_save_theme = await CommandsManager.save_theme_count(task=task_name, theme_count=theme_count)

    if prompts and flag_save_prompts and flag_save_theme:
        await message.answer(f"✅ Промты успешно сгенерированы и сохранены. (<b>{task_name}</b>)")
    else:
        await message.answer(f"❌ Произошла ошибка при генерации или сохранении промтов. "
                             f"Посмотрите лог-файл, чтобы узнать причину. (<b>{task_name}</b>)")

    text, keyboard = await task_prompts(task_name=task_name)
    await message.answer(text, reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()


@router_tasks_prompts.callback_query(lambda call: call.data.startswith('priority-prompts-'), AdminFilter())
async def task_priority_prompts_callback_query(call: CallbackQuery, state: FSMContext):
    task_name = call.data[17:]
    await call.message.edit_text("📊 Отправьте диапазон (от и до включительно) ID приоритетных промтов через дефис.\n\n"
                                 "Чтобы убрать диапазон приоритетных промтов отправьте только дефис.", reply_markup=BACK_TO_TASKS)
    await state.update_data(task_name=task_name)
    await state.set_state(PromptsPriority.message)


@router_tasks_prompts.message(PromptsPriority.message, AdminFilter())
async def get_task_priority_prompts_handler(message: Message, state: FSMContext):
    prompt_data = await state.get_data()
    task_name = prompt_data['task_name']

    if message.text != '-':
        clean_message = message.text.replace(" ", "")

        if not re.fullmatch(r"\d+-\d+", clean_message):
            await message.answer("❌ Пожалуйста, введите диапазон в формате X-Y, где X и Y — числа.", reply_markup=BACK_TO_TASKS)
            return

        start, end = map(int, clean_message.split('-'))
        if start >= end:
            await message.answer("❌ Начало диапазона должно быть меньше конца. Попробуйте снова.", reply_markup=BACK_TO_TASKS)
            return

        text = f"✅ Диапазон {clean_message} приоритетных промтов успешно сохранён. (<b>{task_name}</b>)"
    else:
        text = f"✅ Диапазон приоритетных промтов убран. (<b>{task_name}</b>)"
        clean_message = '-'

    flag = await CommandsManager.save_priority_prompts_to_db(task=task_name, priority_prompts=clean_message)

    if flag:
        await message.answer(text)
    else:
        await message.answer('❌ Произошла ошибка при сохранении диапазона приоритетных промтов.'
                             'Посмотрите лог. (<b>{task_name}</b>)')

    text, keyboard = await task_prompts(task_name=task_name)
    await message.answer(text, reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_prompts.callback_query(lambda call: call.data.startswith('prompts-edit-'), AdminFilter())
async def edit_prompts_callback_query(call: CallbackQuery):
    task_name = call.data[13:]

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Текст промта', callback_data=f'edit-text-prompt-{task_name}')],
        [InlineKeyboardButton(text='Тема промта', callback_data=f'edit-theme-prompt-{task_name}')],
    ])
    await call.message.edit_text("Выберите объект промта для изменения:", reply_markup=keyboard)

@router_tasks_prompts.callback_query(lambda call: call.data.startswith('edit-theme-prompt-'), AdminFilter())
async def edit_theme_prompts_callback_query(call: CallbackQuery, state: FSMContext):
    task_name = call.data[18:]
    await call.message.edit_text("Отправьте ID промта в сообщении.", reply_markup=BACK_TO_TASKS)
    await state.update_data(task_name=task_name)
    await state.set_state(PromptsThemeEdit.message)

@router_tasks_prompts.message(PromptsThemeEdit.message, AdminFilter())
async def get_id_edit_theme_prompts_handler(message: Message, state: FSMContext):
    prompt_data = await state.get_data()
    task_name = prompt_data['task_name']

    if not re.fullmatch(r'\d+', message.text.strip()):
        await message.answer("❌ Укажите корректный ID, состоящий только из цифр.", reply_markup=BACK_TO_TASKS)
        return

    prompt_id = int(message.text.strip())

    connection = sqlite3.connect(DB_DIRECTORY + task_name + '.db', timeout=TIMEOUT_DELAY)
    cursor = connection.cursor()
    cursor.execute("SELECT id FROM Prompts WHERE id = ?", (prompt_id,))
    result = cursor.fetchone()
    connection.close()

    if not result:
        await message.answer("❌ Промт с указанным ID не найден. Убедитесь, что ID указан верно.", reply_markup=BACK_TO_TASKS)
        return

    await message.answer("Отправьте тему промта в сообщении.")
    await state.update_data(task_name=task_name, prompt_id=prompt_id)
    await state.set_state(PromptsThemeEdit.txt)

@router_tasks_prompts.message(PromptsThemeEdit.txt, AdminFilter())
async def edit_theme_prompts_handler(message: Message, state: FSMContext):

    prompt_data = await state.get_data()
    task_name = prompt_data['task_name']
    prompt_id = prompt_data['prompt_id']

    connection = sqlite3.connect(DB_DIRECTORY + task_name + '.db', timeout=TIMEOUT_DELAY)
    cursor = connection.cursor()
    cursor.execute(
        "UPDATE Prompts SET prompt_theme = ? WHERE id = ?",
        (message.text, prompt_id)
    )
    connection.commit()
    connection.close()

    await message.answer("✅ Тема промта успешно обновлена!")
    text, keyboard = await task_prompts(task_name=task_name)
    await message.answer(text, reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_prompts.callback_query(lambda call: call.data.startswith('edit-text-prompt-'), AdminFilter())
async def edit_text_prompts_callback_query(call: CallbackQuery, state: FSMContext):
    task_name = call.data[17:]
    await call.message.edit_text("Отправьте ID промта в сообщении.", reply_markup=BACK_TO_TASKS)
    await state.update_data(task_name=task_name)
    await state.set_state(PromptsEdit.message)

@router_tasks_prompts.message(PromptsEdit.message, AdminFilter())
async def get_id_edit_prompts_handler(message: Message, state: FSMContext):
    prompt_data = await state.get_data()
    task_name = prompt_data['task_name']

    if not re.fullmatch(r'\d+', message.text.strip()):
        await message.answer("❌ Укажите корректный ID, состоящий только из цифр.", reply_markup=BACK_TO_TASKS)
        return

    prompt_id = int(message.text.strip())

    connection = sqlite3.connect(DB_DIRECTORY + task_name + '.db', timeout=TIMEOUT_DELAY)
    cursor = connection.cursor()
    cursor.execute("SELECT id FROM Prompts WHERE id = ?", (prompt_id,))
    result = cursor.fetchone()
    connection.close()

    if not result:
        await message.answer("❌ Промт с указанным ID не найден. Убедитесь, что ID указан верно.", reply_markup=BACK_TO_TASKS)
        return

    await message.answer("Отправьте txt-файл с изменённым текстом промта.")
    await state.update_data(task_name=task_name, prompt_id=prompt_id)
    await state.set_state(PromptsEdit.txt)

@router_tasks_prompts.message(PromptsEdit.txt, AdminFilter())
async def edit_prompts_handler(message: Message, state: FSMContext):

    if not message.document or not message.document.file_name.endswith('.txt'):
        await message.answer("❌ Пожалуйста, отправьте txt-файл.", reply_markup=BACK_TO_TASKS)
        return

    prompt_data = await state.get_data()
    task_name = prompt_data['task_name']
    prompt_id = prompt_data['prompt_id']

    file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
    bot = message.bot
    await bot.download(message.document.file_id, destination=file_path)

    with open(file_path, 'r', encoding='utf-8') as f:
        new_prompt_text = f.read()
    os.remove(file_path)

    connection = sqlite3.connect(DB_DIRECTORY + task_name + '.db', timeout=TIMEOUT_DELAY)
    cursor = connection.cursor()
    cursor.execute(
        "UPDATE Prompts SET prompt = ? WHERE id = ?",
        (new_prompt_text, prompt_id)
    )
    connection.commit()
    connection.close()

    await message.answer("✅ Текст промта успешно обновлён!")
    text, keyboard = await task_prompts(task_name=task_name)
    await message.answer(text, reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_prompts.callback_query(lambda call: call.data.startswith('all-prompts-edit-'), AdminFilter())
async def all_create_prompts_callback_query(call: CallbackQuery, state: FSMContext):
    task_name = call.data[17:]

    await call.message.edit_text("💬 Отправьте xlsx-файл с ID, содержанием промта и его темой.\n\n"
                                 "<b>Формат файла:</b>\n"
                                 "ID | Промт | Тема\n\n"
                                 "<b>Примечание:</b> первая строка считается как название столбцов, данные должны идти со второй строки.", reply_markup=BACK_TO_TASKS
                                 )

    await state.update_data(task_name=task_name)
    await state.set_state(PromptMoreDownload.message)

@router_tasks_prompts.message(PromptMoreDownload.message, AdminFilter())
async def all_create_prompts_handler(message: Message, state: FSMContext):
    data = await state.get_data()
    task_name = data['task_name']

    if message.document:
        if not message.document.file_name.endswith('.xlsx'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .xlsx", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/xlsx/prompts_{uuid.uuid4()}.xlsx"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        connection = sqlite3.connect(DB_DIRECTORY + task_name + '.db', timeout=TIMEOUT_DELAY)
        cursor = connection.cursor()
        wb = load_workbook(file_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            prompt_id, prompt, prompt_theme = row
            if prompt_id:
                update_fields = []
                values = []
                if prompt:
                    update_fields.append("prompt = ?")
                    values.append(prompt)
                if prompt_theme:
                    update_fields.append("prompt_theme = ?")
                    values.append(prompt_theme)
                if update_fields:
                    query = f"UPDATE Prompts SET {', '.join(update_fields)} WHERE id = ?"
                    values.append(prompt_id)
                    cursor.execute(query, tuple(values))
        connection.commit()
        connection.close()

    else:
        await message.answer("❌ Пожалуйста, отправьте xlsx-файл с ID, содержанием промта и его темой.", reply_markup=BACK_TO_TASKS)
        return

    await message.answer('✅ Промты успешно отредактированы!')
    text, keyboard = await task_prompts(task_name=task_name)
    await message.answer(text, reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

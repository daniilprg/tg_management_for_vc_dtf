import asyncio
import os
import re
import sqlite3
import uuid
from itertools import count

import paramiko

import openpyxl
import requests
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery, FSInputFile
from aiogram.filters import CommandStart
from aiogram import Router
from openai import AsyncOpenAI
from openpyxl.reader.excel import load_workbook

from openpyxl.styles import PatternFill

from bot.config import DB_TASK_DIRECTORY, TIMEOUT_DELAY, DB_OPENAI_API_KEY_DIRECTORY, DB_ARTICLES_DIRECTORY, \
    DB_MAIN_ACCOUNTS_DIRECTORY, DB_MULTI_ACCOUNTS_DIRECTORY, DB_XLSX_DIRECTORY
from bot.databases.database_manager import DatabaseManager
from bot.handlers.commands.admins_filter import AdminFilter
from bot.handlers.commands.logging import get_task_logger, log
from bot.handlers.commands.commands_manager import CommandsManager
from bot.handlers.commands.posting_modes.articles_editor import articles_editor_run
from bot.handlers.commands.posting_modes.common import get_accounts
from bot.handlers.commands.posting_modes.main_posting import run_task_script
from bot.handlers.commands.posting_modes.posting_from_db import publishing_db
from bot.handlers.commands.posting_modes.server_posting import server_articles_publishing
from bot.handlers.commands.task_manager import manager
from bot.keyboards.keyboards import tasks_list, get_accounts_from_db, generate_pagination_keyboard

router_tasks_list = Router(name=__name__)

class ApiMessage(StatesGroup):
    message = State()

class LinksCountMessage(StatesGroup):
    message = State()

class ServerMessage(StatesGroup):
    message = State()

class IndexingMessage(StatesGroup):
    message = State()

class Model(StatesGroup):
    message = State()

class PromptTest(StatesGroup):
    message = State()

class PromptImage(StatesGroup):
    message = State()

class TimeoutCycle(StatesGroup):
    message = State()

class TimeoutPost(StatesGroup):
    message = State()

class FlagPost(StatesGroup):
    message = State()

class KeyWord(StatesGroup):
    message = State()

class PostingDB(StatesGroup):
    blacklist_articles = State()

class ManagerArticlesDB(StatesGroup):
    message = State()

class ArticlesEditorProcess(StatesGroup):
    message = State()

class ArticlesEditorParam(StatesGroup):
    message = State()
    message2 = State()

class URLsMoreDownload(StatesGroup):
    message = State()

class PostingServer(StatesGroup):
    account_number = State()
    account_for_publication = State()
    get_txt = State()

tasks_publishing_db = {}
tasks_publishing_server = {}
tasks_articles_editor = {}
tasks = {}

connected_server_data = {
    "Server": {
        "transport": None,
        "sftp": None
    }
}

BACK_TO_TASKS = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
    ])

async def disconnected_from_existing_server() -> None:
    """Отключение от внешнего сервера"""

    if connected_server_data["Server"]["transport"] is not None:
        try:
            connected_server_data["Server"]["transport"].close()
            connected_server_data["Server"]["transport"] = None
            connected_server_data["Server"]["sftp"] = None
        except Exception as e:
            log.debug(f'Произошла ошибка при отключении от внешнего сервера: {e}')

@router_tasks_list.message(CommandStart(), AdminFilter())
async def task_management_handler(message: Message, state: FSMContext):
    """Управление заданиями"""

    await disconnected_from_existing_server()

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'back-to-tasks', AdminFilter())
async def task_management_callback_query(call: CallbackQuery):
    """Управление заданиями"""

    keyboard = await tasks_list()
    await call.message.edit_text("Выберите задание или добавьте новое:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'openai-key', AdminFilter())
async def get_api_key_callback_query(call: CallbackQuery, state: FSMContext):
    """Ожидание API-ключа OpenAI"""

    await call.message.edit_text("Отправьте API-ключ OpenAI.", reply_markup=BACK_TO_TASKS)
    await state.set_state(ApiMessage.message)

@router_tasks_list.message(ApiMessage.message, AdminFilter())
async def upload_api_key_handler(message: Message, state: FSMContext):
    """Загрузка API-ключа OpenAI"""

    with sqlite3.connect(DB_OPENAI_API_KEY_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE ApiKey SET api_key = ? WHERE id = 1",
            (message.text, )
        )
        connection.commit()
    await message.answer("✅ API-ключ успешно загружен.")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'options-connected-server', AdminFilter())
async def get_server_param_callback_query(call: CallbackQuery, state: FSMContext):
    """Ожидание настройки подключения к внешнему серверу"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT host, port, username, password FROM TasksSettings WHERE id = 1")
        host, port, username, password = cursor.fetchone()

    await call.message.edit_text("<b>Текущие параметры подключения</b>:\n"
                                 f"IP: {host}\n"
                                 f"Port: {port}\n"
                                 f"Username: {username}\n"
                                 f"Password: {password}\n\n"
                                 "Отправьте настройки подключения к внешнему серверу.\n\n"
                                 "<b>Формат отправки:</b>\n"
                                 "ip\n"
                                 "port\n"
                                 "username\n"
                                 "password\n\n"
                                 "<b>Примечание:</b> каждый параметр с новой строки", reply_markup=BACK_TO_TASKS)
    await state.set_state(ServerMessage.message)

@router_tasks_list.message(ServerMessage.message, AdminFilter())
async def upload_server_param_handler(message: Message, state: FSMContext):
    """Загрузка настроек подключения к внешнему серверу"""

    try:
        host, port, username, password = message.text.strip().split('\n')
    except:
        await message.answer("❌ Пожалуйста, введите параметры в корректном формате.", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE TasksSettings SET host = ?, port = ?, username = ?, password = ? WHERE id = 1",
            (host, port, username, password)
        )
        connection.commit()

    await message.answer("✅ Настройки подключения к внешнему серверу успешно обновлены.")
    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'task-create', AdminFilter())
async def task_create_callback_query(call: CallbackQuery):
    """Создание задания (Выбор: Основной или Мульти)"""

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Основной', callback_data='main-options-account')],
        [InlineKeyboardButton(text='Мультиаккаунты', callback_data='multi-type')],
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data='back-to-tasks')]
    ])

    await call.message.edit_text("Выберите режим работы:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'main-options-account', AdminFilter())
async def task_main_choice_option_accounts_callback_query(call: CallbackQuery):
    """Выбор режима аккаунтов"""

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Выбрать аккаунт', callback_data='choice-account-for-main')],
        [InlineKeyboardButton(text='Все аккаунты', callback_data='main-type')],
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
    ])

    await call.message.edit_text("Выберите режим работы:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'choice-account-for-main', AdminFilter())
async def task_main_choice_account_callback_query(call: CallbackQuery, state: FSMContext):
    """Выбор аккаунта"""

    accounts = get_accounts_from_db()

    if not accounts:
        await call.message.answer("❌ Нет доступных аккаунтов")

        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

        await state.set_state(None)
        await state.clear()
        return

    keyboard = generate_pagination_keyboard(accounts=accounts, page=0, type_selected='select_account_task')

    await call.message.edit_text("Выберите аккаунт для задания:", reply_markup=keyboard)

    await state.update_data(accounts=accounts,
                            type_selected='select_account_task',
                            page=0)

@router_tasks_list.callback_query(lambda call: call.data.startswith('select_account_task:'), AdminFilter())
async def pagination_select_account_task_callback(call: CallbackQuery, state: FSMContext):
    """Создание задания для одного аккаунта"""

    data = await state.get_data()
    account_index = int(call.data.split(":")[1])
    selected_account = data['accounts'][account_index]
    account_mark = f"{"vc" if "vc" in selected_account[1].lower() else "dtf"}-{selected_account[0]}"

    if account_mark in manager.current_tasks:
        await call.message.edit_text(f"❌ Задание с таким аккаунтом уже создано.")
        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
        return

    task_id = str(uuid.uuid4())
    task = f'Задание-{task_id[:6]}'

    db = DatabaseManager()

    db.create_db_main(task)
    task_type = 'Основной'

    status = 'Ожидание загрузки необходимых ресурсов'
    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            'INSERT INTO Tasks (task_name, task_type, status) VALUES (?, ?, ?)',
            (task, task_type, status)
        )
        connection.commit()

    task_log = get_task_logger(task)
    task_log.debug(f"{task} запущено")

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE Tasks SET last_status = ? WHERE task_name = ?",
            (status, task)
        )
        connection.commit()

    await CommandsManager.update_task_status_db(task=task, status='Приостановлено')

    keyboard = await tasks_list()
    await call.message.edit_text("Выберите задание или добавьте новое:", reply_markup=keyboard)

    start_task = asyncio.create_task(manager.add_task(1,
                                                      run_task_script,
                                                      task,
                                                      [account_mark],
                                                      task,
                                                      task_type,
                                                      call.message.chat.id,
                                                      account_mark))
    tasks[task] = start_task
    await start_task

@router_tasks_list.callback_query(lambda call: call.data == 'main-type', AdminFilter())
@router_tasks_list.callback_query(lambda call: call.data == 'multi-type', AdminFilter())
async def task_with_type_create_callback_query(call: CallbackQuery):
    """Создание задания для нескольких аккаунтов"""

    task_id = str(uuid.uuid4())
    task = f'Задание-{task_id[:6]}'

    db = DatabaseManager()

    accounts = await get_accounts(DB_MAIN_ACCOUNTS_DIRECTORY if call.data == 'main-type'
                                  else DB_MULTI_ACCOUNTS_DIRECTORY)

    accounts_marks = []

    for account in accounts:
        mark = f"{"vc" if "vc" in account[9].lower() else "dtf"}-{account[0]}"
        accounts_marks.append(mark)

        if mark in manager.current_tasks:
            await call.message.edit_text(f"❌ Задание для одного или нескольких таких аккаунтов уже создано.")
            keyboard = await tasks_list()
            await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
            return

    if not accounts:
        await call.message.edit_text(f"❌ Пожалуйста, перед созданием задания загрузите аккаунты.")
        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
        return

    if call.data == 'main-type':
        db.create_db_main(task)
        task_type = 'Основной'
    else:
        db.create_db_multi(task)
        task_type = 'Мультиаккаунты'

    status = 'Ожидание загрузки необходимых ресурсов'
    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            'INSERT INTO Tasks (task_name, task_type, status) VALUES (?, ?, ?)',
            (task, task_type, status)
        )
        connection.commit()

    task_log = get_task_logger(task)
    task_log.debug(f"{task} запущено")

    keyboard = await tasks_list()
    await call.message.edit_text("Выберите задание или добавьте новое:", reply_markup=keyboard)

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE Tasks SET last_status = ? WHERE task_name = ?",
            (status, task)
        )
        connection.commit()

    await CommandsManager.update_task_status_db(task=task, status='Приостановлено')
    
    start_task = asyncio.create_task(manager.add_task(1,
                                                      run_task_script,
                                                      task,
                                                      accounts_marks,
                                                      task,
                                                      task_type,
                                                      call.message.chat.id,
                                                      None))
    tasks[task] = start_task
    await start_task

@router_tasks_list.callback_query(lambda call: call.data == 'test-prompt', AdminFilter())
async def test_prompt_callback_query(call: CallbackQuery, state: FSMContext):
    """Ожидание промта"""

    with sqlite3.connect(DB_OPENAI_API_KEY_DIRECTORY, timeout=TIMEOUT_DELAY) as conn_api:
        cursor_api = conn_api.cursor()
        cursor_api.execute("SELECT api_key FROM ApiKey WHERE id = 1")
        api_key = cursor_api.fetchone()[0]

    if api_key != '-':
        await call.message.edit_text("Отправьте промт в txt-файле.", reply_markup=BACK_TO_TASKS)
        await state.set_state(PromptTest.message)
        await state.update_data(api_key=api_key)
    else:
        await call.message.edit_text("❌ Вы не загрузили API-ключ в базу данных.", reply_markup=BACK_TO_TASKS)

@router_tasks_list.message(PromptTest.message, AdminFilter())
async def test_prompt_handler(message: Message, state: FSMContext):
    """Тестирование промта"""

    if not message.document or not message.document.file_name.endswith('.txt'):
        await message.answer("❌ Пожалуйста, отправьте txt-файл.", reply_markup=BACK_TO_TASKS)
        return

    prompt_data = await state.get_data()
    api_key = prompt_data['api_key']

    file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
    bot = message.bot
    await bot.download(message.document.file_id, destination=file_path)

    with open(file_path, 'r', encoding='utf-8') as f:
        prompt_text = f.read()
    os.remove(file_path)

    with sqlite3.connect(DB_OPENAI_API_KEY_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT model_text FROM ModelAI WHERE id = 1")
        model_text = cursor.fetchone()

    client = AsyncOpenAI(api_key=api_key)
    try:
        chat_completion = await client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": prompt_text,
                }
            ],
            model=model_text,
        )
        chat_response = chat_completion.choices[0].message.content

        if len(chat_response) > 3500:
            file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
            with open(file_path, 'w', encoding='utf-8') as file:
                file.write(chat_response)
            await message.answer_document(FSInputFile(file_path), caption='📜 Ответ слишком длинный, отправка через txt-файл.')
            os.remove(file_path)
        else:
            await message.answer(chat_response)
    except Exception as e:
        await message.answer(f'❌ Ошибка запроса: {str(e)}')

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()


@router_tasks_list.callback_query(lambda call: call.data == 'openai-models', AdminFilter())
async def get_api_models_callback_query(call: CallbackQuery):
    """Модели OpenAI"""

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Модель текста', callback_data=f'text-model'),
         InlineKeyboardButton(text='Модель изображения', callback_data=f'image-model')],
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
    ])

    with sqlite3.connect(DB_OPENAI_API_KEY_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT model_text, model_image FROM ModelAI WHERE id = 1")
        model_text, model_image = cursor.fetchone()

    await call.message.edit_text("Текущие модели:\n"
                                 f"Текст - {model_text}\n"
                                 f"Изображение - {model_image}\n\n"
                                 "Выберите модель для редактирования:",
                                 reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'text-model', AdminFilter())
@router_tasks_list.callback_query(lambda call: call.data == 'image-model', AdminFilter())
async def get_api_text_model_callback_query(call: CallbackQuery, state: FSMContext):
    """Ожидание модели текста и изображения OpenAI"""

    await call.message.edit_text("Отправьте название модели OpenAI в сообщении.", reply_markup=BACK_TO_TASKS)

    await state.update_data(model_type=call.data)
    await state.set_state(Model.message)

@router_tasks_list.message(Model.message, AdminFilter())
async def upload_model_handler(message: Message, state: FSMContext):
    """Загрузка модели"""

    model_data = await state.get_data()
    model_type = model_data['model_type']

    with sqlite3.connect(DB_OPENAI_API_KEY_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        if model_type == 'text-model':
            cursor.execute(
                "UPDATE ModelAI SET model_text = ? WHERE id = 1",
                (message.text, )
            )
            await message.answer("✅ Модель текста по умолчанию успешно изменена.")
        else:
            cursor.execute(
                "UPDATE ModelAI SET model_image = ? WHERE id = 1",
                (message.text, )
            )
            await message.answer("✅ Модель изображения по умолчанию успешно изменена.")
        connection.commit()

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'prompt-image', AdminFilter())
async def get_prompt_image_callback_query(call: CallbackQuery, state: FSMContext):
    """Изменение промта изображения OpenAI"""

    with sqlite3.connect(DB_OPENAI_API_KEY_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT prompt_text FROM PromptImage WHERE id = 1")
        prompt_text = cursor.fetchone()[0]

    await call.message.edit_text("Текущий промт:\n"
                                 f"{prompt_text}\n\n"
                                 "%NAME% - обязательная переменная в промте\n\n"
                                 "Отправьте промт изображения в сообщении.", reply_markup=BACK_TO_TASKS)
    await state.set_state(PromptImage.message)

@router_tasks_list.message(PromptImage.message, AdminFilter())
async def upload_prompt_image_handler(message: Message, state: FSMContext):
    """Загрузка промта изображения"""

    if not '%NAME%' in message.text:
        await message.answer("❌ Промт должен содержать обязательную переменную %NAME%.", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_OPENAI_API_KEY_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE PromptImage SET prompt_text = ? WHERE id = 1",
            (message.text, )
        )
        connection.commit()

    await message.answer("✅ Промт изображения успешно изменён.")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'timeout-cycle', AdminFilter())
async def get_timeout_cycle_callback_query(call: CallbackQuery, state: FSMContext):
    """Изменение задержки цикла"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT timeout_task_cycle FROM TasksSettings WHERE id = 1")
        timeout = cursor.fetchone()[0]

    await call.message.edit_text("Текущая задержка цикла для Основного режима:\n"
                                 f"{int(int(timeout) / 60)} мин.\n\n"
                                 "Отправьте новую задержку (кол-во минут, минимум 1) в сообщении.", reply_markup=BACK_TO_TASKS)
    await state.set_state(TimeoutCycle.message)

@router_tasks_list.message(TimeoutCycle.message, AdminFilter())
async def upload_timeout_cycle_handler(message: Message, state: FSMContext):
    """Загрузка тайм-аута между итерациями"""

    new_timeout = message.text.strip()

    if not re.fullmatch(r'\d+', message.text.strip()):
        await message.answer("❌ Введите корректную задержку цикла.", reply_markup=BACK_TO_TASKS)
        return

    if int(new_timeout) < 1:
        await message.answer("❌ Задержка цикла не может быть меньше одной минуты. Попробуйте снова", reply_markup=BACK_TO_TASKS)
        return

    timeout_in_seconds = int(new_timeout) * 60

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("UPDATE TasksSettings SET timeout_task_cycle = ? WHERE id = 1", (str(timeout_in_seconds),))
        connection.commit()

    await message.answer(f"✅ Задержка цикла по умолчанию успешно обновлена на {new_timeout} мин.")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'timeout-publishing', AdminFilter())
async def get_timeout_publishing_callback_query(call: CallbackQuery, state: FSMContext):
    """Изменение задержки постинга"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT timeout_posting_articles FROM TasksSettings WHERE id = 1")
        timeout = cursor.fetchone()[0]

    await call.message.edit_text("Текущая задержка публикации для Основного режима:\n"
                                 f"{int(int(timeout) / 60)} мин.\n\n"
                                 "Отправьте новую задержку (в минутах) в сообщении.", reply_markup=BACK_TO_TASKS)
    await state.set_state(TimeoutPost.message)

@router_tasks_list.message(TimeoutPost.message, AdminFilter())
async def upload_timeout_publishing_handler(message: Message, state: FSMContext):
    """Загрузка тайм-аута между публикациями"""

    new_timeout = message.text.strip()

    if not re.fullmatch(r'\d+', message.text.strip()):
        await message.answer("❌ Введите корректную задержку постинга.", reply_markup=BACK_TO_TASKS)
        return

    if int(new_timeout) < 1:
        await message.answer("❌ Задержка постинга не может быть меньше одной минуты. Попробуйте снова", reply_markup=BACK_TO_TASKS)
        return

    timeout_in_seconds = int(new_timeout) * 60

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("UPDATE TasksSettings SET timeout_posting_articles = ? WHERE id = 1", (str(timeout_in_seconds),))
        connection.commit()

    await message.answer(f"✅ Задержка постинга статей по умолчанию успешно обновлена на {new_timeout} мин.")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'count-words', AdminFilter())
async def get_count_key_callback_query(call: CallbackQuery, state: FSMContext):
    """Изменение количество загрузки ключевых фраз"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT count_key_words FROM TasksSettings WHERE id = 1")
        count = cursor.fetchone()[0]

    await call.message.edit_text("Текущее количество загрузки ключевых фраз для Основного режима:\n"
                                 f"{count} шт.\n\n"
                                 "Отправьте новое количество загрузки фраз (минимум 1) в сообщении.", reply_markup=BACK_TO_TASKS)
    await state.set_state(KeyWord.message)

@router_tasks_list.message(KeyWord.message, AdminFilter())
async def upload_count_key_handler(message: Message, state: FSMContext):
    """Загрузка количества загружаемых ключей"""

    new_count = message.text.strip()

    if not re.fullmatch(r'\d+', message.text.strip()):
        await message.answer("❌ Введите корректное количество.", reply_markup=BACK_TO_TASKS)
        return

    if int(new_count) < 1:
        await message.answer("❌ Минимальное количество - 1. Попробуйте снова", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("UPDATE TasksSettings SET count_key_words = ? WHERE id = 1", (str(new_count),))
        connection.commit()

    await message.answer(f"✅ Количество загружаемых фраз по умолчанию успешно обновлено на {new_count} шт.")
    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'options-posting', AdminFilter())
async def get_options_callback_query(call: CallbackQuery):
    """Изменение настройки постинга"""

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Для основного режима', callback_data=f'option-main'),
         InlineKeyboardButton(text='Для постинга из БД', callback_data=f'option-db')],
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
    ])

    await call.message.edit_text("Выберите режим настройки:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'option-main', AdminFilter())
async def get_options_main_callback_query(call: CallbackQuery, state: FSMContext):
    """Изменение настройки постинга для Основного режима"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT flag_posting_for_main FROM TasksSettings WHERE id = 1")
        flag = cursor.fetchone()[0]

    await call.message.edit_text("Текущая настройка публикации для Основного режима:\n"
                                 f"{'Публиковать сразу' if flag == 'True' else 'Добавлять в черновики'}\n\n"
                                 "Отправьте новую настройку в сообщении (True - публиковать сразу, False - добавлять в черновики).", reply_markup=BACK_TO_TASKS)
    await state.update_data(type='main')
    await state.set_state(FlagPost.message)

@router_tasks_list.callback_query(lambda call: call.data == 'option-db', AdminFilter())
async def get_options_db_callback_query(call: CallbackQuery, state: FSMContext):
    """Изменение настройки постинга БД"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT flag_posting_db FROM TasksSettings WHERE id = 1")
        flag = cursor.fetchone()[0]

    await call.message.edit_text("Текущая настройка публикации из базы данных:\n"
                                 f"{'Публиковать сразу' if flag == 'True' else 'Добавлять в черновики'}\n\n"
                                 "Отправьте новую настройку в сообщении (True - публиковать сразу, False - добавлять в черновики).", reply_markup=BACK_TO_TASKS)
    await state.update_data(type='db')
    await state.set_state(FlagPost.message)

@router_tasks_list.message(FlagPost.message, AdminFilter())
async def upload_options_db_handler(message: Message, state: FSMContext):
    """Загрузка настройки постинга БД"""

    posting_data = await state.get_data()
    type_posting = posting_data['type']

    new_flag = message.text.strip()

    if new_flag != 'True' and new_flag != 'False':
        await message.answer("❌ Некорректный ввод. Попробуйте снова", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()

        if type_posting == 'db':
            cursor.execute("UPDATE TasksSettings SET flag_posting_db = ? WHERE id = 1", (new_flag,))
        else:
            cursor.execute("UPDATE TasksSettings SET flag_posting_for_main = ? WHERE id = 1", (new_flag,))
        connection.commit()

    await message.answer(f"✅ Настройка постинга {'из базы данных' if type_posting == 'db' 
        else 'для Основного режима'} по умолчанию обновлена: <b>{'Публиковать сразу' if new_flag == 'True' 
        else 'Добавлять в черновики'}</b>"
                         )

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'posting-articles-db', AdminFilter())
async def publishing_db_callback_query(call: CallbackQuery):
    """Публикация статей из базы данных"""

    if 'Task' not in tasks_publishing_db or tasks_publishing_db['Task'].done():

        with sqlite3.connect(DB_ARTICLES_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM Articles")
            articles_count = cursor.fetchone()[0]
            connection.commit()

        keyboard = InlineKeyboardMarkup(inline_keyboard=[])

        if articles_count != 0:
            keyboard.inline_keyboard.append([InlineKeyboardButton(text='🗃 Выгрузка всех статей', callback_data='download-all-articles')])
            keyboard.inline_keyboard.append([InlineKeyboardButton(text='🗑 Удаление статей', callback_data='delete-general-articles')])
            keyboard.inline_keyboard.append([InlineKeyboardButton(text='Запустить постинг', callback_data='engine-publishing-db')])
        else:
            keyboard.inline_keyboard.append(
                [InlineKeyboardButton(text='Статей не обнаружено', callback_data='none')]
            )

        keyboard.inline_keyboard.append(
            [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data='back-to-tasks')]
        )

        await call.message.edit_text("Режим публикации статей из базы данных.\n\n"
                                     "<b>Примечание</b>: параметры публикации берутся с настроек по умолчанию",
                                     reply_markup=keyboard)
    else:
        with sqlite3.connect(DB_ARTICLES_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT status FROM ArticlesStatus WHERE id = 1")
            status = cursor.fetchone()[0]
            connection.commit()

        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='Завершить работу', callback_data=f'publishing-db-stop')],
            [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
        ])
        await call.message.answer("Запущен процесс публикации статей из базы данных.\n\n"
                                  "<b>Статус:</b>\n"
                                  f"{status}",
                                  reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'delete-general-articles', AdminFilter())
async def delete_general_articles_callback_query(call: CallbackQuery, state: FSMContext):
    """Запрос списка статей по ID из общей базы данных"""

    await call.message.edit_text("Отправьте список ID статей, которые нужно удалить из базы данных, в сообщении или txt-файле.\n\n"
                                 "<b>Примечание:</b> каждый ID с новой строки", reply_markup=BACK_TO_TASKS)
    await state.set_state(ManagerArticlesDB.message)

@router_tasks_list.message(ManagerArticlesDB.message, AdminFilter())
async def delete_general_articles_db_handler(message: Message, state: FSMContext):
    """Удаление статей по ID из общей базы данных"""

    if message.document:
        if not message.document.file_name.endswith('.txt'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .txt.", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with open(file_path, 'r', encoding='utf-8') as f:
            list_articles = f.read()
        os.remove(file_path)
    elif message.text:
        list_articles = message.text
    else:
        await message.answer("❌ Пожалуйста, отправьте сообщение или txt-файл.", reply_markup=BACK_TO_TASKS)
        return

    article_ids = [int(article_id.strip()) for article_id in list_articles.split() if article_id.strip().isdigit()]
    if not article_ids:
        await message.answer("❌ Не удалось распознать ID статей.", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_ARTICLES_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()

        cursor.execute("SELECT article_image FROM Articles WHERE id IN ({})".format(
            ','.join('?' * len(article_ids))
        ), article_ids)
        images_to_delete = [row[0] for row in cursor.fetchall()]

        cursor.execute("DELETE FROM Articles WHERE id IN ({})".format(
            ','.join('?' * len(article_ids))
        ), article_ids)
        connection.commit()

    for image_path in images_to_delete:
        if os.path.exists(image_path):
            os.remove(image_path)

    await message.answer(f"✅ Статьи успешно удалены!")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'download-all-articles', AdminFilter())
async def download_all_articles_callback_query(call: CallbackQuery):
    """Выгрузка всех сгенерированных статей"""

    with sqlite3.connect(DB_ARTICLES_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM Articles")
        articles_data = cursor.fetchall()
        connection.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Articles"
    ws.append(["ID", "Text", "Image", "Marks"])

    for row in articles_data:
        ws.append(row)

    file_path = f"bot/assets/xlsx/all_articles.xlsx"
    wb.save(file_path)

    await call.message.answer_document(FSInputFile(file_path),
                                       caption="Конвертированная таблица базы данных в xlsx-файл.")
    os.remove(file_path)

    keyboard = await tasks_list()
    await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'publishing-db-stop', AdminFilter())
async def stop_publishing_from_db_callback_query(call: CallbackQuery):
    """Остановка публикации из общей базы данных статей"""

    if 'Task' in tasks_publishing_db and not tasks_publishing_db['Task'].done():
        tasks_publishing_db['Task'].cancel()
        del tasks_publishing_db['Task']
        await call.message.edit_text(f"Процесс публикации статей из базы данных завершен принудительно.")
    else:
        await call.message.edit_text("Нет активной задачи для остановки.")

    keyboard = await tasks_list()
    await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'engine-publishing-db', AdminFilter())
async def settings_publishing_db_callback_query(call: CallbackQuery, state: FSMContext):
    """Настройка публикации статей из базы данных"""

    accounts = get_accounts_from_db()

    if not accounts:
        await call.message.answer("❌ Нет доступных аккаунтов, на которые можно опубликовать статьи")

        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

        await state.set_state(None)
        await state.clear()
        return

    keyboard = generate_pagination_keyboard(accounts=accounts, page=0, type_selected='select_account_main_posting_database_v1')

    await call.message.edit_text("Выберите аккаунт, на который нужно опубликовать статьи:", reply_markup=keyboard)

    await state.update_data(accounts=accounts,
                            type_selected='select_account_main_posting_database_v1',
                            page=0)

@router_tasks_list.callback_query(lambda call: call.data.startswith('select_account_main_posting_database_v1:'), AdminFilter())
async def pagination_select_account_main_posting_database_v1_callback(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    account_index = int(call.data.split(":")[1])
    selected_account = data['accounts'][account_index]
    account_mark = f"{"vc" if "vc" in selected_account[1].lower() else "dtf"}-{selected_account[0]}"

    accounts = get_accounts_from_db()

    if not accounts:
        await call.message.answer("❌ Нет доступных аккаунтов, на которые можно опубликовать статьи")

        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

        await state.set_state(None)
        await state.clear()
        return

    keyboard = generate_pagination_keyboard(accounts=accounts, page=0, type_selected='select_account_main_posting_database_v2')

    await call.message.edit_text("Выберите аккаунт, с которого нужно брать сгенерированные статьи:", reply_markup=keyboard)

    await state.update_data(account_mark=account_mark,
                            accounts=accounts,
                            type_selected='select_account_main_posting_database_v2',
                            page=0)

@router_tasks_list.callback_query(lambda call: call.data.startswith('select_account_main_posting_database_v2:'), AdminFilter())
async def pagination_select_account_main_posting_database_v2_callback(call: CallbackQuery, state: FSMContext):

    data = await state.get_data()
    account_mark = data['account_mark']

    account_index = int(call.data.split(":")[1])
    selected_account = data['accounts'][account_index]
    source_mark = f"{"vc" if "vc" in selected_account[1].lower() else "dtf"}-{selected_account[0]}"

    await call.message.edit_text("Отправьте чёрный список идентификаторов статей в сообщении или txt-файле.\n\n"
                         "<b>Примечание:</b> каждый ID с новой строки", reply_markup=BACK_TO_TASKS)

    await state.update_data(account_mark=account_mark, source_mark=source_mark)
    await state.set_state(PostingDB.blacklist_articles)

@router_tasks_list.message(PostingDB.blacklist_articles, AdminFilter())
async def start_publishing_db_handler(message: Message, state: FSMContext):
    """Запуск публикации сгенерированных статей из общей базы данных"""

    data = await state.get_data()
    account_mark = data['account_mark']
    source_mark = data['source_mark']

    if message.document:
        if not message.document.file_name.endswith('.txt'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .txt", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with open(file_path, 'r', encoding='utf-8') as f:
            blacklist_articles = f.read()
        os.remove(file_path)
    elif message.text:
        blacklist_articles = message.text
    else:
        await message.answer("❌ Пожалуйста, отправьте сообщение или txt-файл.", reply_markup=BACK_TO_TASKS)
        return

    start_task = asyncio.create_task(manager.add_task(4,
                                                      publishing_db,
                                                      'PostingDB',
                                                      [account_mark],
                                                      account_mark,
                                                      source_mark,
                                                      blacklist_articles,
                                                      message.chat.id,
                                                      ))
    tasks_publishing_db['Task'] = start_task

    await message.answer(f"Запуск процесса публикации общих статей из базы данных.")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()
    await start_task

@router_tasks_list.callback_query(lambda call: call.data == 'posting-articles-server', AdminFilter())
async def posting_articles_server_callback_query(call: CallbackQuery, state: FSMContext):
    """Публикация статей с внешнего сервера"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT host, port, username, password FROM TasksSettings WHERE id = 1")
        host, port, username, password = cursor.fetchone()

    if host == '-' or port == '-' or username == '-' or password == '-':
        await call.message.answer("❌ Пожалуйста, перед публикацией статей с внешнего сервера произведите настройки подключения.")
        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
        return

    if 'Task' not in tasks_publishing_server or tasks_publishing_server['Task'].done():
        try:
            await call.message.edit_text("Подключение к внешнему серверу...")

            transport = paramiko.Transport((host, int(port)))
            transport.connect(username=username, password=password)

            sftp = paramiko.SFTPClient.from_transport(transport)

            connected_server_data["Server"]["transport"] = transport
            connected_server_data["Server"]["sftp"] = sftp

            remote_path = "/root/vc-ru-parser/output/"

            await call.message.edit_text("Получение списка аккаунтов...")

            folders = [entry for entry in sftp.listdir_attr(remote_path) if entry.st_mode & 0o40000]

            await call.message.edit_text(
                "Режим публикации статей с внешнего сервера.\n\n"
                "<b>Примечание</b>: 1) параметры публикации берутся с настроек по умолчанию; 2) Автозамена работает только в том случае, "
                "если настроены параметры в разделе Редактор статей.\n\n"
                "<b>Список доступных аккаунтов внешнего сервера:</b>\n" +
                "\n".join(f"{num}. {folder.filename}" for num, folder in enumerate(folders, start=1)) +
                "\n\nОтправьте номер аккаунта (например, 1) в сообщении.", reply_markup=BACK_TO_TASKS
            )

            await state.update_data(account_folder_list=folders)
            await state.set_state(PostingServer.account_number)

        except Exception as e:
            await call.message.edit_text(f"❌ Произошла ошибка во время подключения к внешнему серверу: {e}", reply_markup=BACK_TO_TASKS)
    else:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='Завершить работу', callback_data='publishing-server-stop')],
            [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
        ])
        await call.message.edit_text("Запущен процесс публикации статей с внешнего сервера.", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'publishing-server-stop', AdminFilter())
async def stop_publishing_from_server_callback_query(call: CallbackQuery):
    """Остановка публикации с внешнего сервера"""

    if 'Task' in tasks_publishing_server and not tasks_publishing_server['Task'].done():
        tasks_publishing_server['Task'].cancel()
        del tasks_publishing_server['Task']
        await call.message.edit_text(f"Процесс публикации статей с внешнего сервера завершен принудительно.")
    else:
        await call.message.edit_text("Нет активной задачи для остановки.")

    keyboard = await tasks_list()
    await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

@router_tasks_list.message(PostingServer.account_number, AdminFilter())
async def get_account_for_posting_articles_server_handler(message: Message, state: FSMContext):
    """
    Получение папки аккаунта, с которого нужно взять статьи.
    Выбор аккаунта, на который нужно опубликовать статьи.
    """

    data = await state.get_data()
    account_folder_list = data['account_folder_list']

    if not message.text.strip().isdigit():
        await message.answer(f"❌ Некорректный ввод номера аккаунта. Попробуйте снова", reply_markup=BACK_TO_TASKS)
        return

    account_index = int(message.text.strip())-1

    if (account_index < 0) or not(0 <= account_index < len(account_folder_list)):
        await message.answer(f"❌ Некорректный ввод номера аккаунта. Попробуйте снова", reply_markup=BACK_TO_TASKS)
        return

    accounts = get_accounts_from_db()

    if not accounts:
        await message.answer("❌ Нет доступных аккаунтов, на которые можно опубликовать статьи")

        keyboard = await tasks_list()
        await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

        await state.set_state(None)
        await state.clear()
        return

    keyboard = generate_pagination_keyboard(accounts=accounts, page=0, type_selected='select_account_server')

    await message.answer("Выберите аккаунт, на который хотите опубликовать статьи:", reply_markup=keyboard)

    await state.update_data(account_with=account_folder_list[account_index].filename,
                            accounts=accounts,
                            type_selected='select_account_server',
                            page=0)

@router_tasks_list.callback_query(lambda call: call.data.startswith('page:'), AdminFilter())
async def pagination_callback(call: CallbackQuery, state: FSMContext):
    """Пагинация списка основных аккаунтов"""

    data = await state.get_data()

    new_page = int(call.data.split(":")[1])
    keyboard = generate_pagination_keyboard(accounts=data['accounts'], page=new_page, type_selected=data['type_selected'])

    await call.message.edit_reply_markup(reply_markup=keyboard)
    await state.update_data(page=new_page)

@router_tasks_list.callback_query(lambda call: call.data.startswith('select_account_server:'), AdminFilter())
async def pagination_select_account_server_callback(call: CallbackQuery, state: FSMContext):
    """Выбор режима публикации статей"""

    data = await state.get_data()
    account_index = int(call.data.split(":")[1])
    selected_account = data['accounts'][account_index]

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='Все статьи', callback_data=f'server-publishing-all-articles')],
            [InlineKeyboardButton(text='Только ID', callback_data=f'server-publishing-only-articles')],
            [InlineKeyboardButton(text='Кроме ID', callback_data=f'server-publishing-without-articles')]
        ])

    await call.message.edit_text(f"Выберите режим публикации:", reply_markup=keyboard)
    await state.update_data(selected_account=selected_account)

@router_tasks_list.callback_query(lambda call: call.data == 'server-publishing-all-articles', AdminFilter())
async def server_publishing_all_articles_callback(call: CallbackQuery, state: FSMContext):
    """Публикация всех статей с внешнего сервера"""

    data = await state.get_data()
    selected_account = data['selected_account']
    account_with = data['account_with']

    account_mark = f"{"vc" if "vc" in selected_account[1].lower() else "dtf"}-{selected_account[0]}"

    sftp = connected_server_data["Server"]["sftp"]

    remote_path = "/root/vc-ru-parser/output/" + account_with + "/"
    local_path = "bot/assets/json/"

    await call.message.edit_text("Получение списка статей...")

    articles_folders = [entry.filename for entry in sftp.listdir_attr(remote_path) if entry.st_mode & 0o40000]
    articles_path = []
    articles_count_download = 0

    await call.message.edit_text(f"Выгрузка статей... ({articles_count_download} из {len(articles_folders)})")

    for folder in articles_folders:
        remote_file = f"{remote_path}{folder}/data.json"
        local_file = os.path.join(local_path, f"{folder}.json")

        try:
            sftp.get(remote_file, local_file)
            articles_path.append(local_file)
            articles_count_download += 1
            await call.message.edit_text(f"Выгрузка статей... ({articles_count_download} из {len(articles_folders)})")
        except FileNotFoundError:
            log.debug(f"Файл {remote_file} не найден.")

    await disconnected_from_existing_server()

    await call.message.edit_text("Запуск процесса постинга статей с внешнего сервера")

    start_task = asyncio.create_task(manager.add_task(5,
                                                      server_articles_publishing,
                                                      'PostingServer',
                                                      [account_mark],
                                                      account_mark,
                                                      articles_path,
                                                      call.message.chat.id
                                                      ))
    tasks_publishing_server['Task'] = start_task

    keyboard = await tasks_list()
    await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()
    await start_task

@router_tasks_list.callback_query(lambda call: call.data == 'server-publishing-only-articles', AdminFilter())
async def server_publishing_only_articles_callback(call: CallbackQuery, state: FSMContext):
    """Публикация статей с определёнными ID"""

    await call.message.edit_text(
        "Отправьте список ID статей, которые вы <b>хотите</b> опубликовать, в сообщении или txt-файле.\n\n"
        "<b>Примечание:</b> каждый ID с новой строки", reply_markup=BACK_TO_TASKS)

    await state.update_data(type='only')
    await state.set_state(PostingServer.get_txt)

@router_tasks_list.callback_query(lambda call: call.data == 'server-publishing-without-articles', AdminFilter())
async def server_publishing_without_articles_callback(call: CallbackQuery, state: FSMContext):
    """Публикация статей без определённых ID"""

    await call.message.edit_text(
        "Отправьте список ID статей, которые вы <b>не хотите</b> опубликовывать, в сообщении или txt-файле.\n\n"
        "<b>Примечание:</b> каждый ID с новой строки", reply_markup=BACK_TO_TASKS)

    await state.update_data(type='without')
    await state.set_state(PostingServer.get_txt)

@router_tasks_list.message(PostingServer.get_txt, AdminFilter())
async def server_publishing_with_type_handler(message: Message, state: FSMContext):
    """Публикация статей с внешнего сервера с определённым типом списка ID"""

    data = await state.get_data()
    selected_account = data['selected_account']
    account_with = data['account_with']
    type_publishing = data['type']

    if message.document:
        if not message.document.file_name.endswith('.txt'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .txt", reply_markup=BACK_TO_TASKS)
            return
        file_path = f"bot/assets/txt/{uuid.uuid4()}.txt"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)
        with open(file_path, 'r', encoding='utf-8') as f:
            list_articles = f.read()
        os.remove(file_path)
    elif message.text:
        list_articles = message.text
    else:
        await message.answer("❌ Пожалуйста, отправьте сообщение или txt-файл.", reply_markup=BACK_TO_TASKS)
        return

    account_mark = f"{"vc" if "vc" in selected_account[1].lower() else "dtf"}-{selected_account[0]}"

    sftp = connected_server_data["Server"]["sftp"]

    remote_path = "/root/vc-ru-parser/output/" + account_with + "/"
    local_path = "bot/assets/json/"

    message = await message.answer("Получение списка статей...")

    articles_folders = [entry.filename for entry in sftp.listdir_attr(remote_path)
                        if entry.st_mode & 0o40000 and ((type_publishing == 'only' and entry.filename in list_articles)
                                                        or (type_publishing != 'only' and not(entry.filename in list_articles)))]
    articles_path = []
    articles_count_download = 0

    message = await message.edit_text(f"Выгрузка статей... ({articles_count_download} из {len(articles_folders)})")

    for folder in articles_folders:
        remote_file = f"{remote_path}{folder}/data.json"
        local_file = os.path.join(local_path, f"{folder}.json")

        try:
            sftp.get(remote_file, local_file)
            articles_path.append(local_file)
            articles_count_download += 1
            await message.edit_text(f"Выгрузка статей... ({articles_count_download} из {len(articles_folders)})")
        except FileNotFoundError:
            log.debug(f"Файл {remote_file} не найден.")

    await disconnected_from_existing_server()

    if articles_path:
        message = await message.edit_text("Запуск процесса постинга статей с внешнего сервера")

        start_task = asyncio.create_task(manager.add_task(4,
                                                          server_articles_publishing,
                                                          'PostingServer',
                                                          [account_mark],
                                                          account_mark,
                                                          articles_path,
                                                          message.chat.id
                                                          ))
        tasks_publishing_server['Task'] = start_task
        keyboard = await tasks_list()
        await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
        await state.set_state(None)
        await state.clear()
        await start_task
    else:
        message = await message.edit_text("Процесс постинга статей с внешнего сервера не может быть запущен.\n\n"
                                          "Найденных статей для публикации: 0 шт.")
        keyboard = await tasks_list()
        await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
        await state.set_state(None)
        await state.clear()


@router_tasks_list.callback_query(lambda call: call.data == 'options-links-indexing', AdminFilter())
async def get_action_indexing_callback_query(call: CallbackQuery):
    """Ожидание настройки индексации"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT indexing FROM TasksSettings WHERE id = 1")
        indexing = cursor.fetchone()[0]

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f'{'✅ Индексация вкл.' if indexing == 'True' else '❌ Индексация выкл.'}', callback_data='default-indexing-trigger'),
         InlineKeyboardButton(text='Параметры индексации', callback_data='edit-param-indexing')],
        [InlineKeyboardButton(text='💳 Остаток на балансе', callback_data='indexing-balance')],
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data='back-to-tasks')]
    ])

    await call.message.edit_text("Выберите действие:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'default-indexing-trigger', AdminFilter())
async def indexing_trigger_callback_query(call: CallbackQuery):
    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT indexing FROM TasksSettings WHERE id = 1")
        indexing_main = cursor.fetchone()[0]

    if indexing_main == 'True':
        indexing = 'False'
    else:
        indexing = 'True'

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE TasksSettings SET indexing = ? WHERE id = 1",
            (indexing, )
        )

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f'{'✅ Индексация вкл.' if indexing == 'True' else '❌ Индексация выкл.'}',
                              callback_data='default-indexing-trigger'),
         InlineKeyboardButton(text='Параметры индексации', callback_data=f'edit-param-indexing')],
        [InlineKeyboardButton(text='💳 Остаток на балансе', callback_data='indexing-balance')],
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
    ])

    await call.message.edit_text("Выберите действие:", reply_markup=keyboard)

@router_tasks_list.message(IndexingMessage.message, AdminFilter())
async def upload_indexing_param_handler(message: Message, state: FSMContext):
    """Загрузка настроек индексации"""

    try:
        api_key, user_id, searchengine, se_type = message.text.strip().split('\n')
    except:
        await message.answer("❌ Пожалуйста, введите параметры в корректном формате.", reply_markup=BACK_TO_TASKS)
        return

    if searchengine != 'google' and searchengine != 'yandex' and searchengine != 'google+yandex':
        await message.answer("❌ Пожалуйста, введите корректный поисковик.", reply_markup=BACK_TO_TASKS)
        return

    if se_type != 'hard' and se_type != 'normal':
        await message.answer("❌ Пожалуйста, введите корректный способ индексации.", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE TasksSettings SET user_id = ?, api_key = ?, searchengine = ?, se_type = ? WHERE id = 1",
            (user_id, api_key, searchengine, se_type)
        )
        connection.commit()

    await message.answer("✅ Настройки индексации успешно обновлены.")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'edit-param-indexing', AdminFilter())
async def get_indexing_callback_query(call: CallbackQuery, state: FSMContext):
    """Ожидание настройки индексации"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT api_key, user_id, searchengine, se_type FROM TasksSettings WHERE id = 1")
        api_key, user_id, searchengine, se_type = cursor.fetchone()

    await call.message.edit_text("<b>Текущие параметры индексации</b>:\n"
                                 f"API-ключ: {api_key}\n"
                                 f"USER_ID: {user_id}\n"
                                 f"Поисковик: {searchengine}\n"
                                 f"Способ индексации: {se_type}\n\n"
                                 "Отправьте настройки индексации.\n"
                                 "Поисковик - google/yandex/google+yandex\n"
                                 "Способ индексации - normal/hard\n\n"
                                 "<b>Формат отправки:</b>\n"
                                 "api-key\n"
                                 "поисковик\n"
                                 "способ индексации\n\n"
                                 "<b>Примечание:</b> каждый параметр с новой строки", reply_markup=BACK_TO_TASKS)
    await state.set_state(IndexingMessage.message)

@router_tasks_list.callback_query(lambda call: call.data == 'indexing-balance', AdminFilter())
async def get_balance_indexing_callback_query(call: CallbackQuery):
    """Остаток на балансе индексации"""

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
    ])

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT api_key, user_id FROM TasksSettings WHERE id = 1")
        api_key, user_id = cursor.fetchone()

    if api_key == '-':
        await call.message.answer("❌ Перед проверкой баланса укажите API-ключ в параметрах индексации.", reply_markup=BACK_TO_TASKS)
        return

    await call.message.edit_text(f'Отправка запроса проверки баланса...')

    try:
        response = requests.get(
            url=f'https://link-indexing-bot.ru/api/users/{user_id}?api_key={api_key}',
        )

        if response.status_code == 200 or response.status_code == 201:
            data = response.json()
            await call.message.edit_text(f'<b>Остаток на балансе</b>: {data['data']['balance']}', reply_markup=keyboard)
            return
        text = (f'❌ Не удалось проверить баланс индексации.\n\n'
                f'<b>Причина</b>: {response.text if response.text else response}')
        await call.message.edit_text(text, reply_markup=keyboard)
    except Exception as e:
        await call.message.edit_text(f'❌ Ошибка при проверке баланса.\n\n'
                                     f'<b>Причина</b>: {e}', reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'articles-editor', AdminFilter())
async def articles_editor_callback_query(call: CallbackQuery):
    """Редактирование статей"""

    if 'Task' not in tasks_articles_editor or tasks_articles_editor['Task'].done():
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='Запустить редактор', callback_data='articles-editor-start')],
            [InlineKeyboardButton(text='Параметры замены', callback_data='articles-editor-param')],
            [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
        ])
        await call.message.edit_text("Режим редактирования статей.", reply_markup=keyboard)
    else:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='Завершить работу', callback_data='articles-editor-stop')],
            [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
        ])
        await call.message.edit_text("Запущен процесс редактирования статей.", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'articles-editor-start', AdminFilter())
async def start_articles_editor_callback_query(call: CallbackQuery, state: FSMContext):
    """Запуск редактора статей"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT currents_replace, new_replace FROM TasksSettings WHERE id = 1")
        currents_replace, new_replace = cursor.fetchone()

    if currents_replace == 'None' or new_replace == 'None':
        await call.message.answer("❌ Параметры замены не указаны")
        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
        return

    accounts = get_accounts_from_db()

    if not accounts:
        await call.message.answer("❌ Нет доступных аккаунтов")

        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

        await state.set_state(None)
        await state.clear()
        return

    keyboard = generate_pagination_keyboard(accounts=accounts, page=0, type_selected='select_account_articles_editor')

    await call.message.edit_text("Выберите аккаунт, на котором нужно отредактировать статьи:", reply_markup=keyboard)
    await state.update_data(accounts=accounts,
                            type_selected='select_account_articles_editor',
                            page=0)

@router_tasks_list.callback_query(lambda call: call.data.startswith('select_account_articles_editor:'), AdminFilter())
async def articles_editor_get_callback(call: CallbackQuery, state: FSMContext):

    data = await state.get_data()
    account_index = int(call.data.split(":")[1])
    selected_account = data['accounts'][account_index]
    account_mark = f"{"vc" if "vc" in selected_account[1].lower() else "dtf"}-{selected_account[0]}"

    await call.message.edit_text(f"Отправьте список ссылок на статей, которые нужно отредактировать, в <b>txt-файле</b>", reply_markup=BACK_TO_TASKS)
    await state.update_data(account_mark=account_mark)
    await state.set_state(ArticlesEditorProcess.message)

@router_tasks_list.message(ArticlesEditorProcess.message, AdminFilter())
async def articles_editor_engine_handler(message: Message, state: FSMContext):
    data = await state.get_data()
    account_mark = data['account_mark']

    if message.document:
        if not message.document.file_name.endswith('.txt'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .txt", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/txt/urls_{uuid.uuid4()}.txt"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

    else:
        await message.answer("❌ Пожалуйста, отправьте txt-файл.", reply_markup=BACK_TO_TASKS)
        return

    start_task = asyncio.create_task(manager.add_task(3,
                                                      articles_editor_run,
                                                      'ArticlesEditor',
                                                      [account_mark],
                                                      account_mark,
                                                      file_path,
                                                      message.chat.id,
                                                      ))
    tasks_articles_editor['Task'] = start_task

    await message.answer(f"Запуск процесса редактора статей.")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()
    await start_task

@router_tasks_list.callback_query(lambda call: call.data == 'articles-editor-stop', AdminFilter())
async def stop_articles_editor_callback_query(call: CallbackQuery):
    """Остановка редактирования статей"""

    if 'Task' in tasks_articles_editor and not tasks_articles_editor['Task'].done():
        tasks_articles_editor['Task'].cancel()
        del tasks_articles_editor['Task']
        await call.message.edit_text(f"Процесс редактирования статей завершен принудительно.")
    else:
        await call.message.edit_text("Нет активной задачи для остановки.")

    keyboard = await tasks_list()
    await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'articles-editor-param', AdminFilter())
async def get_articles_param_callback_query(call: CallbackQuery, state: FSMContext):
    """Ожидание настройки индексации"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT currents_replace, new_replace FROM TasksSettings WHERE id = 1")
        currents_replace, new_replace = cursor.fetchone()

    if currents_replace != 'None':
        currents_replace = currents_replace.strip().split('\n')

    await call.message.edit_text("<b>Текущие параметры замены</b>:\n"
                                 f"Найти: {'не указано' if currents_replace == 'None' else '\n' + '\n'.join(currents_replace)}\n"
                                 f"Заменить на: {'не указано' if new_replace == 'None' else new_replace}\n\n"
                                 "Отправьте параметры, которые нужно <b>найти</b>, в сообщении.\n\n"
                                 "<b>Примечание:</b> каждый параметр с новой строки", reply_markup=BACK_TO_TASKS)
    await state.set_state(ArticlesEditorParam.message)

@router_tasks_list.message(ArticlesEditorParam.message, AdminFilter())
async def edit_param1_articles_editor_handler(message: Message, state: FSMContext):

    if not message.text:
        await message.answer("❌ Пожалуйста, отправьте параметры с новой строки в сообщении.", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE TasksSettings SET currents_replace = ? WHERE id = 1",
            (message.text,)
        )

    await message.answer("Отправьте параметр (один), <b>на который</b> нужно заменить, в сообщении.", reply_markup=BACK_TO_TASKS)
    await state.set_state(ArticlesEditorParam.message2)

@router_tasks_list.message(ArticlesEditorParam.message2, AdminFilter())
async def edit_param2_articles_editor_handler(message: Message, state: FSMContext):

    if not message.text:
        await message.answer("❌ Пожалуйста, отправьте параметр в сообщении.", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE TasksSettings SET new_replace = ? WHERE id = 1",
            (message.text.strip(),)
        )

    await message.answer("✅ Параметры замены успешно обновлены!")
    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'xlsx-urls', AdminFilter())
async def panel_xlsx_urls_callback_query(call: CallbackQuery):

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Выгрузить URLs', callback_data='xlsx-urls-download')],
        [InlineKeyboardButton(text='Массовая загрузка', callback_data='xlsx-urls-upload')],
        [InlineKeyboardButton(text='⬅️ Назад к заданиям', callback_data=f'back-to-tasks')]
    ])
    await call.message.edit_text("Выберите действие:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data == 'xlsx-urls-download', AdminFilter())
async def download_urls_callback_query(call: CallbackQuery):
    with sqlite3.connect(DB_XLSX_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT keys, urls_accounts FROM Xlsx")
        rows = cursor.fetchall()

    if not rows:
        await call.message.answer("❌ Нет данных для выгрузки.")
        keyboard = await tasks_list()
        await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "URLs"

    green_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")

    for keys_str, urls_accounts in rows:

        keys_list = keys_str.split("\n")

        if urls_accounts is None:
            for i, key in enumerate(keys_list):
                ws.append([key])

                if i == 0:
                    ws[f"A{ws.max_row}"].fill = green_fill
            continue

        urls_accounts_list = [entry.partition(" ")[::2] for entry in urls_accounts.split(" | ")]

        for i, key in enumerate(keys_list):
            row_data = [key]

            for url, acc in urls_accounts_list:
                row_data.extend([url, acc])

            ws.append(row_data)

            if i == 0:
                ws[f"A{ws.max_row}"].fill = green_fill

    file_path = f"bot/assets/xlsx/urls_{uuid.uuid4()}.xlsx"
    wb.save(file_path)

    await call.message.answer_document(FSInputFile(file_path),
                                       caption="Конвертированная таблица базы данных в xlsx-файл.")
    os.remove(file_path)

    keyboard = await tasks_list()
    await call.message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

@router_tasks_list.callback_query(lambda call: call.data.startswith('xlsx-urls-upload'), AdminFilter())
async def all_create_urls_callback_query(call: CallbackQuery, state: FSMContext):
    await call.message.edit_text("💬 Отправьте xlsx-файл с ключами.\n\n"
                                 "<b>Формат файла:</b>\n"
                                 "ID | Ключи\n\n"
                                 "<b>Примечание:</b> первая строка считается как название столбцов, данные должны идти со второй строки. В одной ячейке xlsx-файла должны хранится тема и её ключи. Если такой ID темы и её ключей уже существует в базе данных, то произойдёт замена, иначе добавление.",
                                 reply_markup=BACK_TO_TASKS
                                )
    await state.set_state(URLsMoreDownload.message)

@router_tasks_list.message(URLsMoreDownload.message, AdminFilter())
async def all_create_urls_handler(message: Message, state: FSMContext):
    if message.document:
        if not message.document.file_name.endswith('.xlsx'):
            await message.answer("❌ Пожалуйста, отправьте файл с расширением .xlsx", reply_markup=BACK_TO_TASKS)
            return

        file_path = f"bot/assets/xlsx/urls_{uuid.uuid4()}.xlsx"
        bot = message.bot
        await bot.download(message.document.file_id, destination=file_path)

        with sqlite3.connect(DB_XLSX_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
            cursor = connection.cursor()
            wb = load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                keys_id, keys = row

                cursor.execute("SELECT COUNT(*) FROM Xlsx WHERE id = ?", (keys_id,))
                exists = cursor.fetchone()[0] > 0

                if exists:
                    cursor.execute("UPDATE Xlsx SET keys = ? WHERE id = ?", (str(keys), keys_id))
                else:
                    cursor.execute("INSERT INTO Xlsx (id, keys) VALUES (?, ?)", (keys_id, str(keys)))

            connection.commit()

    else:
        await message.answer("❌ Пожалуйста, отправьте xlsx-файл с названием шаблонов и их содержимым.", reply_markup=BACK_TO_TASKS)
        return

    keyboard = await tasks_list()
    await message.answer('✅ Ключи успешно выгружены и добавлены!')
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)
    await state.set_state(None)
    await state.clear()

@router_tasks_list.callback_query(lambda call: call.data == 'options-auto-links', AdminFilter())
async def get_articles_links_count_callback_query(call: CallbackQuery, state: FSMContext):
    """Ожидание параметра количество ссылок"""

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT articles_links_count FROM TasksSettings WHERE id = 1")
        count = cursor.fetchone()[0]

    await call.message.edit_text("<b>Текущие параметры:</b>\n"
                                 f"Количество отправляемых ссылок: {count}"
                                 "\n\nОтправьте количество ссылок, которое нужно добавить в статью.", reply_markup=BACK_TO_TASKS)
    await state.set_state(LinksCountMessage.message)

@router_tasks_list.message(LinksCountMessage.message, AdminFilter())
async def upload_articles_links_count_handler(message: Message, state: FSMContext):
    """Загрузка параметра количество ссылок"""

    count = message.text.strip()

    if not count.isdigit():
        await message.answer("❌ Пожалуйста, отправьте число.", reply_markup=BACK_TO_TASKS)
        return

    with sqlite3.connect(DB_TASK_DIRECTORY, timeout=TIMEOUT_DELAY) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE TasksSettings SET articles_links_count = ? WHERE id = 1",
            (count, )
        )
        connection.commit()
    await message.answer("✅ Количество ссылок для Автоперелинковки успешно обновлено.")

    keyboard = await tasks_list()
    await message.answer("Выберите задание или добавьте новое:", reply_markup=keyboard)

    await state.set_state(None)
    await state.clear()